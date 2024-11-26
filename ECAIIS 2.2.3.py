# ECAIIS（EC Automated Inventory Input System）
# EC在庫表自動入力システム
#####  #####  #####  #  #  #####
#      #      #   #  #  #  #
#####  #      #####  #  #  #####
#      #      #   #  #  #      #
#####  #####  #   #  #  #  #####

from __future__ import print_function

# プログラム情報----------------------------------------------------------------
PROGRAM_NAME = "ECAIIS" # プログラム名
PROGRAM_VERSION = "2.2.3" # バージョン
PROGRAM_UPDATE = "2024年10月8日" # 更新日
#----------------------------------------------------------------

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException,WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
import time
import datetime
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import PySimpleGUI as sg
import shutil
import pandas as pd
import numpy as np
import os
import re
import csv
import os.path
import io
import chardet
import xlwings as xw
import webbrowser
import glob
import calendar

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload,MediaFileUpload

from pydrive.drive import GoogleDrive
from pydrive.auth import GoogleAuth
import psutil

from cryptography.fernet import Fernet
from io import StringIO
import configparser

import pygetwindow as gw
import pyautogui

import openpyxl

# 全ての実行中のプロセスを取得
for proc in psutil.process_iter():
    try:
        # プロセス名を取得
        process_name = proc.name()
        # プロセス名が'Excel'である場合（Windowsでは'EXCEL.EXE'、macOSでは'Microsoft Excel'）
        if process_name == 'EXCEL.EXE' or process_name == 'Microsoft Excel':
            # プロセスを終了
            proc.kill()
    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
        pass

today = datetime.datetime.now()

# 現在のPCユーザー名を取得
PC_username = os.getlogin()

# スクリプトは読み取り専用のメタデータアクセスを要求するスコープを定義
SCOPES = ['https://www.googleapis.com/auth/drive']

def decryption(plain_filename,key_filename):
    """ファイルの復号を行なう

    Args:
        plain_filename (str): 暗号化ファイルのリンク先
        key_filename (str): 鍵のリンク先

    Returns:
        str: 復号したデータ
    """
    # 鍵をファイルから読み込む
    with open(f'{key_filename}', 'rb') as my_key_file:
        key = my_key_file.read()
    f = Fernet(key)

    # 暗号化したファイルを読み込む
    with open(f'{plain_filename}', 'rb') as file:
        text = file.read()
    # 読み込んだbyte型の暗号データを複合化
    plain_text = f.decrypt(text).decode('utf-8')
    return plain_text

def get_config_value(data, section, key):
    """ ini形式の文字列から指定データを取得

    Args:
        data (str): データ
        section (str): セクション名
        key (str): キー名

    Returns:
        str: 指定データ
    """
    # StringIOオブジェクトを作成
    data_io = StringIO(data)

    # configparserオブジェクトを作成
    config = configparser.ConfigParser()
    config.read_file(data_io)

    # 指定したセクションの指定したキーのデータを取得
    value = config.get(section, key)

    return value

def close_window(windows_name:list):
    """nameを含むwindowを閉じる

    Args:
        windows_name (list): 閉じるウィンドウ名一覧
    """
    # 全てのウィンドウのタイトルを取得
    all_windows = gw.getWindowsWithTitle('')

    for name in windows_name:
        # Chromeウィンドウを検索
        chrome_windows = [window for window in all_windows if name in window.title.lower()]
        print(f"{name}:{chrome_windows}")

        # Chromeウィンドウが見つかった場合、それらを閉じる
        if len(chrome_windows) > 0:
            choice = sg.popup_ok_cancel(f'現在開いている{name}を強制的に閉じます。\n'
                                '問題がない場合は,"OK"を押してください。\n'
                                '保存が必要な場合、"NG"を選択し、各ソフト上で保存を行なってから、再度ECAIISを起動してください。')
            # Check the user's selection and act accordingly
            if choice == 'OK':
                for window in chrome_windows:
                    # ウィンドウをアクティブにする
                    window.activate()
                    # ウィンドウを閉じる
                    pyautogui.hotkey('ctrl', 'w')
            else:
                # Terminate the program
                sg.popup('ECAIISを終了します。')
                exit()

def remove_file(name_list:list,folder_path:str):
    """ファイルを削除する関数

    Args:
        name_list (list): 削除したいファイルに含む名前リスト
        folder_path (str): フォルダパス
    """
    for filename in os.listdir(folder_path):
        for name in name_list:
            if name in filename:
                file_path = os.path.join(folder_path, filename)
                try:
                    os.remove(file_path)
                    print(f'{filename} を削除しました。')
                except Exception as e:
                    print(f'{filename} の削除中にエラーが発生しました: {e}')

def version_matching_confirmation():
    """chromedriver と chrome(browser)のバージョンの互換性を確認する
        最新版のchromedriverのダウンロードサイト
        https://googlechromelabs.github.io/chrome-for-testing/
    """
    log(["chromedriver と chrome(browser)のバージョンが一致しているか確認中..."])
    # 最新のchromedriverをインストール
    driver_path = ChromeDriverManager().install()
    if os.path.splitext(driver_path)[1] != '.exe':
        driver_dir_path = os.path.dirname(driver_path)
        driver_path = os.path.join(driver_dir_path, 'chromedriver.exe')
    # 仮想環境へコピー
    shutil.copy2(driver_path,'.venv/Lib/site-packages/chromedriver_binary/chromedriver.exe')
    # chromedriverインスタンスの設定
    chromedriver = webdriver.Chrome(service=Service(driver_path))
    # chromedriver のバージョン
    chromedriver_version = chromedriver.capabilities['chrome']['chromedriverVersion'].split(' ')[0]
    # chrome(browser) のバージョン
    chromebrowser_version = chromedriver.capabilities['browserVersion'].split(' ')[0]
    # chromedriverインスタンスを閉じる
    chromedriver.quit()

    if chromedriver_version.split('.')[0] == chromebrowser_version.split('.')[0] \
        and chromedriver_version.split('.')[1] == chromebrowser_version.split('.')[1] \
            and chromedriver_version.split('.')[2] == chromebrowser_version.split('.')[2]:
                log(["バージョンが一致しました。"])
    else:
        log(["バージョンが一致していません。"])
        log([f"chromedriver_version:{chromedriver_version}"])
        log([f"chromebrowser_version{chromebrowser_version}"])

def download_for_googledrive(saving_foldername:str,search_list:list,folder_id:str = None):
    """Google WorkSpace API を用いてGoogle Driveの共有フォルダからファイルをダウンロードする。

    Args:
        saving_foldername (str): 保存先
        search_list (list): 保存条件

    Returns:
        _type_: _description_
    """
    download_files=[]
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('conf/token.json'):
        creds = Credentials.from_authorized_user_file('conf/token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'conf/credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # 新たに取得した認証情報をtoken.jsonに保存する
        with open('conf/token.json', 'w') as token:
            token.write(creds.to_json())

    # Google DriveのAPIサービスを作成し、ユーザーがアクセスできる最初の10ファイルのリストを取得する。
    # エラーが発生した場合は、エラーメッセージを出力します。
    try:
        service = build('drive', 'v3', credentials=creds)
        if folder_id:
            query = f"'{folder_id}' in parents and ("
        else:
            query = "("
        for index, criteria in enumerate(search_list):
            query += f'name contains "{criteria}"'
            if index != len(search_list) - 1:
                query += ' and '
        query += ') and trashed = false'

        # Call the Drive v3 API
        results = service.files().list(
                pageSize=100,
                fields="nextPageToken, files(id, name)",
                q=query
        ).execute()
        items = results.get('files', [])

        if not items:
            print('No files found.')
            return
        print('Download Files:')
        for item in items:
            if not item["name"] in download_files:
                download_files.append(item['name'])
                print(u'{0} ({1})'.format(item['name'], item['id']))
                # ダウンロードする
                request = service.files().get_media(fileId = item['id'])
                file_path = os.path.join(saving_foldername,item['name'])
                file = io.FileIO(file_path, mode='wb')
                downloader = MediaIoBaseDownload(file,request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()

    except HttpError as error:
        # TODO(developer) - Handle errors from drive API.
        print(f'An error occurred: {error}')
    except PermissionError:
        print(f"Permission denied for file: {file_path}")

    return download_files

def upload_to_googledrive(folder_id:str,search_folder_name_list:list,upload_folder_name:str,upload_file_name_list:list):
    """Google WorkSpace API を用いてGoogle Driveの共有フォルダにファイルをアップロードする。

    Args:
        folder_id (str): 共有フォルダのID
        search_folder_name_list (list): アップロードフォルダ先指定
        upload_file_name_list (list): アップロードするファイル名
    """
    #!Googleドライブの認証
    #Googleサービスを認証
    gauth = GoogleAuth()
    #資格情報ロードするか、存在しない場合は空の資格情報を作成
    gauth.LoadCredentialsFile("conf/mycreds.txt")

    #初回のみ(OAuth 2.0 クライアント IDのjsonからアクセス権取得)
    #Googleサービスの資格情報がない場合
    if gauth.credentials is None:
        #ユーザーから認証コードを自動的に受信しローカルWebサーバーを設定
        gauth.LocalWebserverAuth()
    #アクセストークンが存在しないか、期限切れかの場合
    elif gauth.access_token_expired:
        #Googleサービスを認証をリフレッシュする
        gauth.Refresh()
    #どちらにも一致しない場合
    else:
        #Googleサービスを承認する
        gauth.Authorize()
    #資格情報をtxt形式でファイルに保存する
    gauth.SaveCredentialsFile("conf/mycreds.txt")

    #Googleドライブの認証処理
    drive = GoogleDrive(gauth)

    #! 共有フォルダへのアップロード
    #アップロード元フォルダパス指定
    path = os.path.join(upload_folder_name)

    # googledrive共有フォルダのEC在庫表のフォルダID
    FOLDER_ID = folder_id
    # 指定した共有フォルダ下のファイルやフォルダを取得
    file_list = drive.ListFile({'q': f"'{FOLDER_ID}' in parents and trashed=false"}).GetList()
    # アップロードするフォルダ先を指定する
    for f in file_list:
        applicable = True
        for search_name in search_folder_name_list:
            if not search_name in f["title"]:
                applicable = False
                break
        if applicable == True:
            FOLDER_ID = f["id"]

    # dataフォルダにあるファイルを全てアップロード
    for x in os.listdir(path):
        # 検索ファイル名リストに該当するファイルがあるかどうかを確認する
        applicable = True
        for search_name in upload_file_name_list:
            if not search_name in x:
                applicable = False
                break
        # 該当するファイル名がある場合
        if applicable == True:
            # アップロードするファイル名
            FILE_NAME = x

            # 共有フォルダ内にすでにアップロードされているファイルのIDを取得
            file_list = drive.ListFile({'q': f"'{FOLDER_ID}' in parents and trashed=false and title='{FILE_NAME}'"}).GetList()

            if len(file_list) > 0:
                # 既存のファイルが存在する場合は、最初のファイルを上書き
                file = file_list[0]
                file.SetContentFile(f"{path}/{FILE_NAME}")  # 'new_content.txt'はあなたがアップロードしたい新しいファイルのパスです
                file.Upload()
            else:
                # 既存のファイルが存在しない場合は、新しいファイルを作成
                file = drive.CreateFile({'title': FILE_NAME, 'parents': [{'id': FOLDER_ID}]})
                file.SetContentFile(f"{path}/{FILE_NAME}")  # 'new_content.txt'はあなたがアップロードしたい新しいファイルのパスです
                file.Upload()

class Download:
    """各ECサイトから注文レポートをダウンロードするAPI
    """
    def __init__(self):
        pass

    def amazon(self,login_id,login_pw):
        """ amazon から 注文レポートをダウンロードする。\n
            (未使用関数)

        Args:
            login_id (str): ログインID
            login_pw (str): ログインPW
        """
        #! オプション設定
        options = Options()

        # プロファイル設定
        # profile_iniからプロファイル情報を取得
        config = configparser.ConfigParser()
        config.read('conf/profile.ini')
        PROFILE_DIR = config.get('GoogleChromeProfile','user-data-dir')
        PROFILE_DIRECTORY = config.get('GoogleChromeProfile','profile-directory')
        options.add_argument('--user-data-dir=' + PROFILE_DIR)
        options.add_argument('--profile-directory=' + PROFILE_DIRECTORY)

        # 仮想環境下のchromedriver.exeファイルを実行する
        webdriver_service = Service('.venv/Lib/site-packages/chromedriver_binary/chromedriver.exe')
        driver = webdriver.Chrome(service=webdriver_service, options=options)

        driver.maximize_window()
        driver.execute_script("document.body.style.zoom = '100%'")
        url = 'https://sellercentral-japan.amazon.com/order-reports-and-feeds/reports/ref=xx_orderrpt_dnav_xx'
        driver.get(url) # MyCiSページを開く

        try:
            ### 共通ID
            # username
            elem = driver.find_element(By.ID,"ap_email")
            elem.send_keys(Keys.CONTROL + "a")
            elem.send_keys(Keys.DELETE)
            elem.send_keys(login_id)
            # password
            elem = driver.find_element(By.ID,"ap_password")
            elem.send_keys(Keys.CONTROL + "a")
            elem.send_keys(Keys.DELETE)
            elem.send_keys(login_pw)
            # button(次へ)
            elem = driver.find_element(By.ID,"signInSubmit")
            elem.click()
        except NoSuchElementException:
            print("要素が見つかりませんでした。")

        time.sleep(3)

        try:
            driver.close()
        except WebDriverException as e:
            print(e)

    def rakuten(self,common_id,common_pw,personal_id,personal_pw,download_id,download_pw):
        """rakuten から 注文レポートをダウンロードする。

        Args:
            common_id (_type_): 共通ID
            common_pw (_type_): 共通PW
            personal_id (_type_): 個人ID
            personal_pw (_type_): 個人PW
            download_id (_type_): ダウンロードID
            download_pw (_type_): ダウンロードPW
        """
        #! オプション設定
        options = Options()

        # プロファイル設定
        # profile_iniからプロファイル情報を取得
        config = configparser.ConfigParser()
        config.read('conf/profile.ini')
        PROFILE_DIR = config.get('GoogleChromeProfile','user-data-dir')
        PROFILE_DIRECTORY = config.get('GoogleChromeProfile','profile-directory')
        options.add_argument('--user-data-dir=' + PROFILE_DIR)
        options.add_argument('--profile-directory=' + PROFILE_DIRECTORY)

        # 仮想環境下のchromedriver.exeファイルを実行する
        webdriver_service = Service('.venv/Lib/site-packages/chromedriver_binary/chromedriver.exe')
        driver = webdriver.Chrome(service=webdriver_service, options=options)

        driver.maximize_window()
        driver.execute_script("document.body.style.zoom = '100%'")
        url = 'https://glogin.rms.rakuten.co.jp/?sp_id=1'
        driver.get(url) # MyCiSページを開く

        ### 共通ID
        # username
        elem = driver.find_element(By.ID,"rlogin-username-ja")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(common_id)
        # password
        elem = driver.find_element(By.ID,"rlogin-password-ja")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(common_pw)
        # button(次へ)
        elem = driver.find_element(By.NAME,"submit")
        elem.click()

        time.sleep(3)
        ### 個人ID
        #username
        elem = driver.find_element(By.ID,"rlogin-username-2-ja")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(personal_id)
        # password
        elem = driver.find_element(By.ID,"rlogin-password-2-ja")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(personal_pw)
        # button(次へ)
        elem = driver.find_element(By.NAME,"submit")
        elem.click()

        # button
        elem = driver.find_element(By.NAME,"submit")
        elem.click()
        # button
        elem = driver.find_element(By.CSS_SELECTOR,".btn-reset")
        elem.click()

        #todo --- 2024/7/10 追加 -----------------------------------------------

        time.sleep(3)
        # 注意喚起メッセージを閉じる
        try:
            elem = driver.find_element(By.CLASS_NAME, "rf-grid-column")
            elem = driver.find_element(By.XPATH, '//*[@id="message-ja"]/section/div/div/div/div/button')
            elem.click()
        except:
            pass

        time.sleep(3)
        # お知らせを閉じる
        try:
            elem = driver.find_element(By.CLASS_NAME, "tac")
            elem = driver.find_element(By.XPATH, '//*[@id="confirm"]/p/button')
            elem.click()
        except:
            pass

        #todo end----------------------------------------------------------------

        time.sleep(3)
        # お知らせを閉じる
        try:
            elem = driver.find_element(By.CLASS_NAME, "eccMessagePopTitle")
            elem = driver.find_element(By.XPATH, '//*[@id="overContainer1"]/div[2]/div/a')
            elem.click()
        except:
            pass

        # 受注・決済管理
        elem = driver.find_element(By.ID,"com_gnavi0200")
        elem.click()
        # 受注管理
        elem = driver.find_element(By.ID,"com_gnavi0201")
        elem.click()

        # 受注データダウンロード
        elem = driver.find_element(By.ID,"mm_sub0201_19")
        elem.click()

        # 通常購入データ
        elem = driver.find_element(By.CSS_SELECTOR,'[href="/rms/mall/csvdl/CD02_01_001?dataType=opp_order"]')
        elem.click()

        ### 期間指定
        start_day = dates_list[0] + relativedelta(days= -1)
        end_day = dates_list[-1]
        # 開始日選択
        # 日付
        dropdown = driver.find_element(By.NAME,"fromYmd")
        select = Select(dropdown)
        select.select_by_value(f"{start_day.year}-{str(start_day.month).zfill(2)}-{str(start_day.day).zfill(2)}")
        # 時
        dropdown = driver.find_element(By.NAME,"fromH")
        select = Select(dropdown)
        select.select_by_value("0")
        # 分
        dropdown = driver.find_element(By.NAME,"fromM")
        select = Select(dropdown)
        select.select_by_value("0")
        # 終了日選択
        # 日付
        dropdown = driver.find_element(By.NAME,"toYmd")
        select = Select(dropdown)
        select.select_by_value(f"{end_day.year}-{str(end_day.month).zfill(2)}-{str(end_day.day).zfill(2)}")
        # 時
        dropdown = driver.find_element(By.NAME,"toH")
        select = Select(dropdown)
        select.select_by_value("23")
        # 分
        dropdown = driver.find_element(By.NAME,"toM")
        select = Select(dropdown)
        select.select_by_value("59")

        # データを作成する
        elem = driver.find_element(By.ID,"dataCreateBtn")
        elem.click()

        # 前：ファイル一覧を取得する
        folder_path = f"C:/Users/{PC_username}/Downloads"
        file_list_before = os.listdir(folder_path)

        try:
            elem = driver.find_element(By.XPATH, "//*[contains(text(),'この条件でのデータ件数は0件です。')]")
            print("該当データなし")
        except NoSuchElementException:
            while True:
                try:
                    # CSVデータダウンロード時のユーザ名、パスワード入力
                    #username
                    elem = driver.find_element(By.ID,"user")
                    elem.send_keys(Keys.CONTROL + "a")
                    elem.send_keys(Keys.DELETE)
                    elem.send_keys(download_id)
                    # password
                    elem = driver.find_element(By.ID,"passwd")
                    elem.send_keys(Keys.CONTROL + "a")
                    elem.send_keys(Keys.DELETE)
                    elem.send_keys(download_pw)
                    # データをダウンロードする
                    elem = driver.find_element(By.ID,"downloadBtn")
                    elem.click()
                    break
                except:
                    time.sleep(5)

            time.sleep(10)
            # 後：ファイル一覧を取得する
            file_list_after = os.listdir(folder_path)

            new_files = list(set(file_list_after) - set(file_list_before))
            if new_files:
                new_file = new_files[0]
            else:
                print("新たにダウンロードされたファイルはありません")

            while True:
                try:
                    # コピー元ファイルパス
                    source = f"C:/Users/{PC_username}/Downloads/{new_file}"
                    # コピー先ファイルパス
                    destination = f"temp/rakuten_{date.year}{str(date.month).zfill(2)}.csv"
                    # ファイルのコピー
                    shutil.copy(source, destination)
                    time.sleep(5)
                    break
                except Exception as e:
                    print(e)
                    time.sleep(5)
                    pass

        try:
            driver.close()
        except WebDriverException as e:
            print(e)

    def yahoo(self,login_id,login_pw):
        """yahoo から 注文データをダウンロードする。

        Args:
            login_id (_type_): ログインID
            login_pw (_type_): ログインPW
        """
        #! オプション設定
        options = Options()
        # プロファイル設定
        # profile_iniからプロファイル情報を取得
        config = configparser.ConfigParser()
        config.read('conf/profile.ini')
        PROFILE_DIR = config.get('GoogleChromeProfile','user-data-dir')
        PROFILE_DIRECTORY = config.get('GoogleChromeProfile','profile-directory')
        options.add_argument('--user-data-dir=' + PROFILE_DIR)
        options.add_argument('--profile-directory=' + PROFILE_DIRECTORY)

        # Webドライバインスタンス生成
        # 仮想環境下のchromedriver.exeファイルを実行する
        webdriver_service = Service('.venv/Lib/site-packages/chromedriver_binary/chromedriver.exe')
        driver = webdriver.Chrome(service=webdriver_service, options=options)
        # ウィンドウ最大化とズーム設定
        driver.maximize_window()
        driver.execute_script("document.body.style.zoom = '100%'")
        url = 'https://login.yahoo.co.jp/config/login?.src=bizmgr&.lg=jp&.intl=jp&.suppreg_skip=1&.done=https%3A%2F%2Flogin.bizmanager.yahoo.co.jp%2Fyidlogin%3F.pass%3D0%26.done%3Dhttps%253A%252F%252Fpro.store.yahoo.co.jp%252Fpro.taiyo-rebirth%252Forder%252Fmanage%252Findex%26.src%3Dnone'
        driver.get(url) # MyCiSページを開く

        ### ログイン
        # id
        try:
            elem = driver.find_element(By.ID,"login_handle")
            time.sleep(1)
            elem.send_keys(Keys.CONTROL + "a")
            elem.send_keys(Keys.DELETE)
            elem.send_keys(login_id)
            # button(次へ)
            elem = driver.find_element(By.XPATH,"//button[text()='次へ']")
            elem.click()
        except NoSuchElementException:
            pass
        except:
            pass

        try:
            # ログイン → パスワード入力 → ログイン
            # 既にパスワードが入力されている場合を想定し、ログインを先に押す
            # button(次へ)
            time.sleep(5)
            elem = driver.find_element(By.XPATH,"//button[text()='ログイン']")
            elem.click()
            time.sleep(1)
            # password
            elem = driver.find_element(By.ID,"password")
            time.sleep(1)
            elem.send_keys(Keys.CONTROL + "a")
            elem.send_keys(Keys.DELETE)
            elem.send_keys(login_pw)
            # button(次へ)
            time.sleep(5)
            elem = driver.find_element(By.XPATH,"//button[text()='ログイン']")
            elem.click()
        except NoSuchElementException:
            pass

        ### 日時設定
        start_day = dates_list[0] + relativedelta(days = -1)
        end_day = dates_list[-1]
        # 開始日選択
        # 日付
        elem = driver.find_element(By.ID,"OrderTimeFromDayE")
        time.sleep(3)
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(f"{start_day.year}/{str(start_day.month).zfill(2)}/{str(start_day.day).zfill(2)}")
        # 時
        dropdown = driver.find_element(By.NAME,"OrderTimeFromHourE")
        select = Select(dropdown)
        select.select_by_value("0000")
        # 終了日時選択
        # 日付
        elem = driver.find_element(By.ID,"OrderTimeToDayE")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(f"{end_day.year}/{str(end_day.month).zfill(2)}/{str(end_day.day).zfill(2)}")
        # 時
        dropdown = driver.find_element(By.NAME,"OrderTimeToHourE")
        select = Select(dropdown)
        select.select_by_value("2359")

        # 注文データダウンロード
        elem = driver.find_element(By.XPATH,'//img[@src="https://s.yimg.jp/images/storecreatorpro/img/1.0.0/common/ic_download.png"]/..')
        elem.click()
        download_dir = f"C:/Users/{PC_username}/Downloads/"

        # 商品系データダウンロード
        for i in range(1,11):
            try:
                if i == 1:
                    elems = driver.find_elements(By.LINK_TEXT,f"{i}")
                else:
                    elems = driver.find_elements(By.LINK_TEXT,f"1～{i}")

                for elem in elems:
                    # ファイルダウンロード前のファイルリスト
                    before_files = os.listdir(download_dir)
                    elem.click()
                    time.sleep(10)

                    after_files = os.listdir(download_dir)

                    # 新たに追加されたファイルを探す
                    new_files = list(set(after_files) - set(before_files))
                    if new_files:
                        new_file = new_files[0]
                    else:
                        print("新たにダウンロードされたファイルはありません")

                    while True:
                        try:
                            # コピー元ファイルパス
                            source = f"{download_dir}{new_file}"
                            # コピー先ファイルパス
                            if "date" in new_file:
                                destination = f"temp/yahoo_{date.year}{str(date.month).zfill(2)}.csv"
                            elif "shohinjohou" in new_file:
                                destination = f"temp/yahoo_{date.year}{str(date.month).zfill(2)}_shohinjohou.csv"
                            else:
                                break
                            # ファイルのコピー
                            shutil.copy(source, destination)
                            time.sleep(5)
                            break
                        except Exception as e:
                            time.sleep(5)
                            print(e)
                            pass
            except NoSuchElementException:
                pass
            except IndexError:
                break

        try:
            driver.close()
        except WebDriverException as e:
            print(e)

    def amazon_cvr(self,login_id,login_pw):
        """ amazon から 注文レポートをダウンロードする。\n
            (未使用関数)

        Args:
            login_id (str): ログインID
            login_pw (str): ログインPW
        """
        #! オプション設定
        options = Options()

        # プロファイル設定
        # profile_iniからプロファイル情報を取得
        config = configparser.ConfigParser()
        config.read('conf/profile.ini')
        PROFILE_DIR = config.get('GoogleChromeProfile','user-data-dir')
        PROFILE_DIRECTORY = config.get('GoogleChromeProfile','profile-directory')
        options.add_argument('--user-data-dir=' + PROFILE_DIR)
        options.add_argument('--profile-directory=' + PROFILE_DIRECTORY)

        # 仮想環境下のchromedriver.exeファイルを実行する
        webdriver_service = Service('.venv/Lib/site-packages/chromedriver_binary/chromedriver.exe')
        driver = webdriver.Chrome(service=webdriver_service, options=options)

        driver.maximize_window()
        driver.execute_script("document.body.style.zoom = '100%'")

        url = "https://sellercentral-japan.amazon.com/business-reports/ref=xx_sitemetric_dnav_xx#/dashboard"
        driver.get(url)

        try:
            ### 共通ID
            # username
            elem = driver.find_element(By.ID,"ap_email")
            elem.send_keys(Keys.CONTROL + "a")
            elem.send_keys(Keys.DELETE)
            elem.send_keys(login_id)
            # password
            elem = driver.find_element(By.ID,"ap_password")
            elem.send_keys(Keys.CONTROL + "a")
            elem.send_keys(Keys.DELETE)
            elem.send_keys(login_pw)
            # button(次へ)
            elem = driver.find_element(By.ID,"signInSubmit")
            elem.click()
        except NoSuchElementException:
            print("要素が見つかりませんでした。")

        # 指定日
        selected_date = datetime.datetime(values['selectdate_year'],values['selectdate_month'],values['selectdate_day'])

        # 一年前の日付を計算
        start_date = selected_date - timedelta(days=365)

        # 各月の1日を表示するためのリストを作成
        dates = [start_date.replace(day=1)]

        # 今日までの各月の1日をリストに追加
        while dates[-1] < selected_date:
            # 次の月の1日を計算
            next_month = dates[-1].month % 12 + 1
            next_year = dates[-1].year + (dates[-1].month // 12)
            next_month_date = dates[-1].replace(year=next_year, month=next_month, day=1)

            # リストに追加
            dates.append(next_month_date)

        # 最後の日付が今日を超えていたら削除
        if dates[-1] > selected_date :
            dates.pop()

        for date in dates:
            end_day = date.replace(day=calendar.monthrange(date.year, date.month)[1])
            if end_day.year == selected_date.year and end_day.month == selected_date.month and end_day.day > selected_date.day:
                end_day = selected_date
            url = f'https://sellercentral-japan.amazon.com/business-reports/ref=xx_sitemetric_dnav_xx#/report?id=102%3ADetailSalesTrafficByChildItem&chartCols=&columns=0%2F1%2F2%2F3%2F8%2F9%2F14%2F15%2F20%2F21%2F26%2F27%2F28%2F29%2F30%2F31%2F32%2F33%2F34%2F35%2F36%2F37&fromDate={date.year}-{str(date.month).zfill(2)}-{str(date.day).zfill(2)}&toDate={date.year}-{str(date.month).zfill(2)}-{str(end_day.day).zfill(2)}'
            driver.get(url) # MyCiSページを開く
            time.sleep(1)
            driver.refresh()
            time.sleep(3)

            download_dir = f"C:/Users/{PC_username}/Downloads/"
            before_files = os.listdir(download_dir)

            elem = driver.find_elements(By.CLASS_NAME,'css-1lafdix')
            elem[0].click()
            time.sleep(3)

            after_files = os.listdir(download_dir)
            # 新たに追加されたファイルを探す
            new_files = list(set(after_files) - set(before_files))
            if new_files:
                new_file = new_files[0]
            else:
                print("新たにダウンロードされたファイルはありません")

            while True:
                try:
                    # コピー元ファイルパス
                    source = f"{download_dir}{new_file}"
                    # コピー先ファイルパス
                    if "BusinessReport" in new_file:
                        destination = f"temp/amazoncvr_{date.year}{str(date.month).zfill(2)}.csv"
                    else:
                        break
                    # ファイルのコピー
                    shutil.copy(source, destination)
                    time.sleep(1)
                    break
                except Exception as e:
                    time.sleep(1)
                    print(e)
                    pass

        try:
            driver.close()
        except WebDriverException as e:
            print(e)

class FileFormatConversion:
    """ファイル形式を変換
    """
    def __init__(self) -> None:
        self.filename = ""
        self.output_filename = ""
        pass

    def txt_to_csv(self):
        with open(self.filename, 'rb') as f:
            result = chardet.detect(f.read())
            encoding = result['encoding']

        # 推定したエンコーディングでファイルを開く
        try:
            df = pd.read_csv(self.filename, delimiter='\t', encoding=encoding)
        except UnicodeDecodeError as e:
            print("Error occurred at position:", e.start, "to", e.end)

        # UTF-8でエンコードされたCSVファイルとして出力する
        df.to_csv("temp/"+self.output_filename, index=False, encoding=encoding)

def setlist_excel(filename,sheet_name,header_skip=False):
    """ 指定ファイル読み込み、リストに格納し、リストを返す。
        ヘッダースキップ

    Args:
        filename (_type_): 読み込むファイル名

    Returns:
        _type_: リスト
    """
    LIST=[]

    start_row = 2 if header_skip is True else 1

    file_name = f"{filename}"
    workbook = openpyxl.load_workbook(file_name,data_only=True)
    worksheet = workbook[sheet_name]

    #シートの最終行を取得
    Sheet_Max_Row = worksheet.max_row
    #シートの最終列を取得
    Sheet_Max_Clm = worksheet.max_column

    for i in range(start_row ,Sheet_Max_Row + 1):
        LIST0 = []
        for j in range(1,Sheet_Max_Clm + 1):
            LIST0.append(worksheet.cell(row=i, column=j).value)
        LIST.append(LIST0)

    return LIST

def save_excel(filename,data):
    try:
        workbook = openpyxl.Workbook()
        workbook.save(filename)

        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook["Sheet"]
        for i in range(1,len(data)+1):
            for j in range(1,len(data[0])+1):
                worksheet.cell(row=i,column=j,value=data[i-1][j-1])

        workbook.save(filename)

    except Exception as e:
        print(e)

class Preparation:
    """ データ準備（リスト化）
    """
    def __init__(self):
        pass

    def amazon_issue(self):
        """Amazon出庫数取得

        Returns:
            _type_: _description_
        """
        amazon_issue_list = np.zeros((len(skulist),31))
        amazonFBA_issue_list = np.zeros((len(skulist),31))
        try:
            with open(f"temp/amazon_{selected_date.year}{str(selected_date.month).zfill(2)}.csv",mode = "r") as f:
                reader = csv.reader(f)
                next(reader) # ヘッダー行をスキップ
                for row in reader:
                    # 注文ステータスが「SHipped（出荷済み）」もしくは「Pending（出荷待ち）」の場合、
                    if row[4] == "Shipped" or "Pending":
                        t0 = re.split("[/:T+-]",row[2])
                        # 日付分け
                        for date in dates_list:
                            yesterday = date + datetime.timedelta(days= -1)
                            if (int(t0[0]) == yesterday.year and int(t0[1]) == yesterday.month and int(t0[2]) == yesterday.day and  int(t0[3]) >= 12) \
                                or (int(t0[0]) == date.year and int(t0[1]) == date.month and int(t0[2]) == date.day and  int(t0[3]) < 12):
                                # skulistと照合し、amazon_issue_listの入力位置を決める
                                for i in range(len(skulist)):
                                    # sku一致
                                    if row[11] == skulist[i][5]:
                                        if row[5] == "Merchant":
                                            amazon_issue_list[i,date.day - 1] = int(amazon_issue_list[i,date.day - 1]) + int(row[17])
                                        elif row[5] == "Amazon":
                                            amazonFBA_issue_list[i,date.day - 1] = int(amazonFBA_issue_list[i,date.day - 1]) + int(row[17])
        except FileNotFoundError:
            print("No file found.")

        return amazon_issue_list,amazonFBA_issue_list

    def rakuten_issue(self):
        """楽天出庫数取得

        Returns:
            _type_: _description_
        """
        rakuten_issue_list = np.zeros((len(skulist),31))
        try:
            with open(f"temp/rakuten_{selected_date.year}{str(selected_date.month).zfill(2)}.csv",mode = "r") as f:
                reader = csv.reader(f)
                next(reader) # ヘッダー行をスキップ
                for row in reader:
                    t0 = re.split("[- /:]",row[4])
                    # 日付分け
                    for date in dates_list:
                        yesterday = date + datetime.timedelta(days= -1)
                        if (int(t0[0]) == yesterday.year and int(t0[1]) == yesterday.month and int(t0[2]) == yesterday.day and  int(t0[3]) >= 12) \
                            or (int(t0[0]) == date.year and int(t0[1]) == date.month and int(t0[2]) == date.day and  int(t0[3]) < 12):
                            # skulistと照合し、rakuten_issue_listの入力位置を決める
                            for i in range(len(skulist)):
                                # sku一致
                                if row[156] == "" or row[156] is None:
                                    if row[155] == skulist[i][7]:
                                        rakuten_issue_list[i,date.day - 1] = int(rakuten_issue_list[i,date.day - 1]) + int(row[76])
                                else:
                                    if row[156] == skulist[i][7]:
                                        rakuten_issue_list[i,date.day - 1] = int(rakuten_issue_list[i,date.day - 1]) + int(row[76])

        except FileNotFoundError:
            print("No file found.")

        return rakuten_issue_list

    def yahoo_issue(self):
        """Yahoo出庫数

        Returns:
            _type_: _description_
        """
        order_data = [] # [注文日時,オーダーID]
        yahoo_issue_list = np.zeros((len(skulist),31))

        # 商品情報からオーダー情報を取得
        try:
            with open(f"temp/yahoo_{selected_date.year}{str(selected_date.month).zfill(2)}_shohinjohou.csv",mode = "r") as f:
                reader = csv.reader(f)
                next(reader) # ヘッダー行をスキップ
                for row in reader:
                    order_data.append(row)
        except FileNotFoundError:
            print("No file found.")

        # 取得したオーダー情報と注文日情報から出庫数データを作成
        try:
            with open(f"temp/yahoo_{selected_date.year}{str(selected_date.month).zfill(2)}.csv",mode = "r") as f:
                reader = csv.reader(f)
                next(reader) # ヘッダー行をスキップ
                for row in reader:
                    t0 = re.split("[- /:]",row[1])
                    # 日付分け
                    for date in dates_list:
                        yesterday = date + datetime.timedelta(days= -1)
                        if (int(t0[0]) == yesterday.year and int(t0[1]) == yesterday.month and int(t0[2]) == yesterday.day and  int(t0[3]) >= 12) \
                            or (int(t0[0]) == date.year and int(t0[1]) == date.month and int(t0[2]) == date.day and  int(t0[3]) < 12):
                            # オーダー情報を探索
                            for order_datum in order_data:
                                # オーダーIDが一致
                                if row[0] == order_datum[0]:
                                    # skulistと照合し、rakuten_issue_listの入力位置を決める
                                    for i in range(len(skulist)):
                                        # sku一致した
                                        if order_datum[3] == skulist[i][9]:
                                            yahoo_issue_list[i,date.day - 1] = int(yahoo_issue_list[i,date.day - 1]) + int(order_datum[2])
        except FileNotFoundError:
            print("No file found.")

        return yahoo_issue_list

    def returns(self):
        """返品数取得
        Returns:
            _type_: _description_
        """

        amazon_return_list = np.zeros((len(skulist),31))
        amazonFBA_return_list = np.zeros((len(skulist),31))
        rakuten_return_list = np.zeros((len(skulist),31))
        yahoo_return_list = np.zeros((len(skulist),31))

        try:
            file_name = f'temp\\返品_{selected_date.year}-{str(selected_date.month).zfill(2)}.xlsx'
            workbook = openpyxl.load_workbook(file_name,data_only=True)
            worksheet = workbook['返品数入力']

            for row in range(2,1000):
                t0 = worksheet['B' + str(row)].value
                if t0 == "" or t0 is None:
                    break
                else:
                    t1 = re.split("[- /:]", t0.strftime("%Y/%m/%d"))
                    for date in dates_list:
                        if int(t1[0]) == date.year and int(t1[1]) == date.month and int(t1[2]) == date.day:
                            for i in range(len(skulist)):
                                if worksheet['I' + str(row)].value == skulist[i][5] or worksheet['I' + str(row)].value ==skulist[i][7] or worksheet['I' + str(row)].value == skulist[i][9]:
                                    if worksheet['C' + str(row)].value == "amazon":
                                        amazon_return_list[i][date.day - 1] = amazon_return_list[i][date.day - 1] + int(worksheet['E' + str(row)].value)
                                    elif worksheet['C' + str(row)].value == "amazonFBA":
                                        amazonFBA_return_list[i][date.day - 1] = amazonFBA_return_list[i][date.day - 1] + int(worksheet['E' + str(row)].value)
                                    elif worksheet['C' + str(row)].value == "rakuten":
                                        rakuten_return_list[i][date.day - 1] = rakuten_return_list[i][date.day - 1] + int(worksheet['E' + str(row)].value)
                                    elif worksheet['C' + str(row)].value == "yahoo":
                                        yahoo_return_list[i][date.day - 1] = yahoo_return_list[i][date.day - 1] + int(worksheet['E' + str(row)].value)
        except FileNotFoundError:
            print("No file found.")
        return amazon_return_list,amazonFBA_return_list,rakuten_return_list,yahoo_return_list

    def sample(self):
        """サンプル数
        Returns : [商品名、数量、シリーズ]
        """
        LIST = []

        try:
            file_name = f"temp\\サンプル・不良・調整_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}.xlsx"
            workbook = openpyxl.load_workbook(file_name,data_only=True)
            worksheet = workbook['入力']

            for row in range(2,1000):
                t0 = worksheet['B' + str(row)].value
                if t0 == "" or t0 is None:
                    break
                else:
                    content_name = worksheet['C' + str(row)].value
                    if content_name == "サンプル":
                        LIST.append([worksheet['D' + str(row)].value,\
                                    int(worksheet['E' + str(row)].value),\
                                    worksheet['F' + str(row)].value
                                    ])

        except FileNotFoundError:
            print("No file found.")

        return LIST

    def defective(self):
        """不良数
        Returns : [商品名、数量、シリーズ]
        """
        LIST = []

        try:
            file_name = f"temp\\サンプル・不良・調整_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}.xlsx"
            workbook = openpyxl.load_workbook(file_name,data_only=True)
            worksheet = workbook['入力']

            for row in range(2,1000):
                t0 = worksheet['B' + str(row)].value
                if t0 == "" or t0 is None:
                    break
                else:
                    content_name = worksheet['C' + str(row)].value
                    if content_name == "不良":
                        LIST.append([worksheet['D' + str(row)].value,\
                                    int(worksheet['E' + str(row)].value),\
                                    worksheet['F' + str(row)].value
                                    ])

        except FileNotFoundError:
            print("No file found.")

        return LIST


    def adjust(self):
        """調整数
        Returns : [商品名、数量、シリーズ]
        """
        LIST = []

        try:
            file_name = f"temp\\サンプル・不良・調整_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}.xlsx"
            workbook = openpyxl.load_workbook(file_name,data_only=True)
            worksheet = workbook['入力']

            for row in range(2,1000):
                t0 = worksheet['B' + str(row)].value
                if t0 == "" or t0 is None:
                    break
                else:
                    content_name = worksheet['C' + str(row)].value
                    if content_name == "調整":
                        LIST.append([worksheet['D' + str(row)].value,\
                                    int(worksheet['E' + str(row)].value),\
                                    worksheet['F' + str(row)].value
                                    ])

        except FileNotFoundError:
            print("No file found.")

        return LIST

    def orders(self):
        """ 塗屋への発注数取得
        """
        LIST = []
        try:
            file_name = f'temp\\発注_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}.xlsx'
            workbook = openpyxl.load_workbook(file_name,data_only=True)
            worksheet = workbook['発注数入力']

            for row in range(2,1000):
                t0 = worksheet['B' + str(row)].value
                t1 = worksheet['C' + str(row)].value
                if t0 == "" or t0 is None:
                    break
                else:
                    t2 = re.split("[- /:]", t0.strftime("%Y/%m/%d"))
                    if int(t2[0]) == selected_date.year and int(t2[1]) == selected_date.month:
                        LIST.append([t0.strftime("%Y/%m/%d"),\
                                    t1.strftime("%Y/%m/%d"),\
                                    worksheet['D' + str(row)].value,\
                                    worksheet['E' + str(row)].value,\
                                    worksheet['F' + str(row)].value,\
                                    worksheet['G' + str(row)].value
                                    ])

        except FileNotFoundError:
            print("No file found.")

        return LIST

    def receipts(self):
        """ 塗屋からの入庫数取得
        """
        LIST = []
        for filename in glob.glob("temp/*"):
            if f"山家_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}" in filename:
                try:
                    with open(f"{filename}","r") as f:
                        reader = csv.reader(f)
                        next(reader)
                        for row in reader:
                            if row[1] == "":
                                break
                            else:
                                LIST.append(row)

                except Exception as e:
                    print(e)
                    pass

        return LIST

    def deliveries(self):
        """ FBA納品数取得

        return:
            [納品日、商品名、納品数、sku]
        """
        deliveries_list = []

        try:
            file_name = f'temp\\FBA納品_{selected_date.year}-{str(selected_date.month).zfill(2)}.xlsx'
            workbook = openpyxl.load_workbook(file_name,data_only=True)
            worksheet = workbook['入力']

            for row in range(2,1000):
                t0 = worksheet['B' + str(row)].value
                if t0 == "" or t0 is None:
                    break
                else:
                    t1 = re.split("[- /:]", t0.strftime("%Y/%m/%d"))
                    if int(t1[0]) == selected_date.year and int(t1[1]) == selected_date.month:
                        deliveries_list.append([t0.strftime("%Y/%m/%d"),\
                                                worksheet['C' + str(row)].value,\
                                                worksheet['D' + str(row)].value,\
                                                worksheet['E' + str(row)].value,\
                                                worksheet['G' + str(row)].value
                                            ])

        except FileNotFoundError:
            print("No file found.")

        return deliveries_list

    def orderRest(self):
        """注残数取得
        """
        LIST = []
        try:
            file_name = f'data\\【新EC在庫表】{series_name}_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx'

            workbook = openpyxl.load_workbook(file_name,read_only=True,data_only=True)
            worksheet = workbook['発注・入庫']

            for row in range(7,1000):
                product_name = worksheet['A' + str(row)].value
                if product_name == "" or product_name is None:
                    break
                else:
                    try:
                        order_rest_value = int(worksheet['AO' + str(row)].value) # 注残数
                        if not order_rest_value == 0:
                            color = worksheet['B' + str(row)].value
                            LIST.append([product_name, series_name, color, order_rest_value])
                    except ValueError:
                        pass

        except FileNotFoundError:
            print("No file found.")

        return LIST

    def lastMonthOrderRest(self):
        """「（自動更新）注残」から注残数取得
        """
        LIST = []
        try:
            file_name = f"temp\\（自動更新）注残数_{str(one_month_ago.year)}-{str(one_month_ago.month).zfill(2)}.xlsx"

            workbook = openpyxl.load_workbook(file_name,read_only=True,data_only=True)
            worksheet = workbook["Sheet1"]

            for i in range(1,1000):
                name = worksheet['A' + str(i)].value
                if name == "" or name is None:
                    break
                else:
                    try:
                        series = worksheet['B' + str(i)].value
                        color = worksheet['C' + str(i)].value
                        order_rest_value = worksheet['D' + str(i)].value
                        LIST.append([name,series,color,order_rest_value])
                    except ValueError:
                        pass

        except FileNotFoundError:
            print("No file found.")

        return LIST

    def orderQuantity(self,List):
        """注文数自動算出のためのデータ準備

        Args:
            List (_type_): _description_

        Returns:
            [品名、シリーズ、色、現在庫、適正在庫数、発注時入数、注文残り]
        """
        try:
            file_name = f'data\\【新EC在庫表】{series_name}_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx'

            workbook = openpyxl.load_workbook(file_name,read_only=True,data_only=True)
            worksheet = workbook['総在庫表']

            for row in range(4,1000):
                if worksheet['B' + str(row)].value == "" or worksheet['B' + str(row)].value is None:
                    break
                else:
                    try:
                        value1 = worksheet['I' + str(row)].value
                        value2 = worksheet['J' + str(row)].value
                        if float(value2) <= 1.5:
                            product_name = worksheet['C' + str(row)].value
                            color = worksheet['D' + str(row)].value
                            data3 = float(worksheet['E' + str(row)].value) # 現在庫
                            data4 = float(worksheet['H' + str(row)].value) # 適正在庫数
                            worksheet1 = workbook['発注・入庫']
                            data5 = float(worksheet1['C' + str(row+3)].value) # 発注時入数
                            data6 = float(worksheet1['AO' + str(row+3)].value) # 注文残り
                            List.append([product_name, series_name, color, data3, data4, data5, data6])
                    except ValueError:
                        pass

        except FileNotFoundError:
            print("No file found.")

        return List

class Input:
    """ ダウンロードしたファイルからデータを取得
    """
    def __init__(self):
        self.amazon_issue_list = [] # amazon出庫数リスト
        self.amazonFBA_issue_list = [] # amazonFBA出庫数リスト
        self.rakuten_issue_list = [] # rakuten出庫数リスト
        self.yahoo_issue_list = [] # yahoo出庫数リスト
        self.amazon_return_list = [] # amazon返品数リスト
        self.amazonFBA_return_list = [] # amazon返品数リスト
        self.rakuten_return_list = [] # amazon返品数リスト
        self.yahoo_return_list = [] # amazon返品数リスト

    def single(self):
        ws = wb.sheets['単品出庫表']

        # B列の最後のセルの行数を一度だけ計算
        last_row_in_B = ws.range('B' + str(ws.cells.last_cell.row)).end('up').row
        # B列の全値を一度だけ取得
        b_values = ws.range(f'B1:B{last_row_in_B}').value

        for i in range(len(skulist)):
            if skulist[i][2] == series_name and skulist[i][3] == "単品":
                if skulist[i][5] in b_values:
                    row = b_values.index(skulist[i][5]) + 1
                    for j in range(31):
                    # 既に取得したB列の値を使用
                        column = 8 + j * 6

                        if self.amazon_issue_list[i][j] > 0:
                            ws.cells(row, column).value = self.amazon_issue_list[i][j]
                        if self.amazon_return_list[i][j] > 0:
                            ws.cells(row, column + 1,).value = self.amazon_return_list[i][j]
                        if self.rakuten_issue_list[i][j] > 0:
                            ws.cells(row, column + 2).value = self.rakuten_issue_list[i][j]
                        if self.rakuten_return_list[i][j] > 0:
                            ws.cells(row, column + 3).value = self.rakuten_return_list[i][j]
                        if self.yahoo_issue_list[i][j] > 0:
                            ws.cells(row, column + 4).value = self.yahoo_issue_list[i][j]
                        if self.yahoo_return_list[i][j] > 0:
                            ws.cells(row, column + 5).value = self.yahoo_return_list[i][j]

    def set(self):
        ws = wb.sheets['セット商品出庫表']

        # B列の最後のセルの行数を一度だけ計算
        last_row_in_B = ws.range('B' + str(ws.cells.last_cell.row)).end('up').row
        # B列の全値を一度だけ取得
        b_values = ws.range(f'B1:B{last_row_in_B}').value

        for i in range(len(skulist)):
            if skulist[i][2] == series_name and skulist[i][3] == "セット":
                if skulist[i][5] in b_values:
                    row = b_values.index(skulist[i][5]) + 1
                    for j in range(31):
                    # 既に取得したB列の値を使用
                        column = 15 + j * 6

                        if self.amazon_issue_list[i][j] > 0:
                            ws.cells(row, column).value = self.amazon_issue_list[i][j]
                        if self.amazon_return_list[i][j] > 0:
                            ws.cells(row, column + 1).value = self.amazon_return_list[i][j]
                        if self.rakuten_issue_list[i][j] > 0:
                            ws.cells(row, column + 2).value = self.rakuten_issue_list[i][j]
                        if self.rakuten_return_list[i][j] > 0:
                            ws.cells(row, column + 3).value = self.rakuten_return_list[i][j]
                        if self.yahoo_issue_list[i][j] > 0:
                            ws.cells(row, column + 4).value = self.yahoo_issue_list[i][j]
                        if self.yahoo_return_list[i][j] > 0:
                            ws.cells(row, column + 5).value = self.yahoo_return_list[i][j]

    def FBA(self):
        ws = wb.sheets['セット商品FBA出庫表']

        # B列の最後のセルの行数を一度だけ計算
        last_row_in_B = ws.range('B' + str(ws.cells.last_cell.row)).end('up').row
        # B列の全値を一度だけ取得
        b_values = ws.range(f'B1:B{last_row_in_B}').value

        for i in range(len(skulist)):
            if skulist[i][2] == series_name and skulist[i][3] == "セット":
                if skulist[i][5] in b_values:
                    row = b_values.index(skulist[i][5]) + 1
                    for j in range(31):
                    # 既に取得したB列の値を使用
                        column = 15 + j * 2

                        if self.amazonFBA_issue_list[i][j] > 0:
                            ws.cells(row, column).value = self.amazonFBA_issue_list[i][j]
                        if self.amazonFBA_return_list[i][j] > 0:
                            ws.cells(row, column + 1).value = self.amazonFBA_return_list[i][j]

    def sample(self,sample_list):
        """サンプル数入力
        """
        filter_list = [row for row in sample_list if row[2] == series_name]

        ws = wb.sheets['自社倉庫在庫表']
        #! 初期化
        for row in range(4, 104):  # R to AAに対応
            ws.cells(row, 13).value = 0

        # 入力
        for row in filter_list:
            for i in range(4,104):
                product_name = ws.cells(i,3).value
                if product_name == "" or product_name is None:
                    break
                else:
                    if product_name == row[0]:
                        value = int(ws.cells(i,13).value)
                        # 既に入力されているデータがある場合は、加算。
                        if value == 0:
                            ws.cells(i,13).value = int(row[1])
                        else:
                            ws.cells(i,13).value = value + int(row[1])

    def defective(self,defective_list):
        """不良数入力
        """
        filter_list = [row for row in defective_list if row[2] == series_name]

        ws = wb.sheets['自社倉庫在庫表']
        #! 初期化
        for row in range(4, 104):  # R to AAに対応
            ws.cells(row, 12).value = 0

        # 入力
        for row in filter_list:
            for i in range(4,104):
                product_name = ws.cells(i,3).value
                if product_name == "" or product_name is None:
                    break
                else:
                    if product_name == row[0]:
                        value = int(ws.cells(i,12).value)
                        # 既に入力されているデータがある場合は、加算。
                        if value == 0:
                            ws.cells(i,12).value = int(row[1])
                        else:
                            ws.cells(i,12).value = value + int(row[1])

    def adjust(self,adjust_list):
        """調整数入力
        """
        filter_list = [row for row in adjust_list if row[2] == series_name]

        ws = wb.sheets['自社倉庫在庫表']
        #! 初期化
        for row in range(4, 104):  # R to AAに対応
            ws.cells(row, 18).value = 0

        # 入力
        for row in filter_list:
            for i in range(4,104):
                product_name = ws.cells(i,3).value
                if product_name == "" or product_name is None:
                    break
                else:
                    if product_name == row[0]:
                        value = int(ws.cells(i,18).value)
                        # 既に入力されているデータがある場合は、加算。
                        if value == 0:
                            ws.cells(i,18).value = int(row[1])
                        else:
                            ws.cells(i,18).value = value + int(row[1])

        ws = wb.sheets['FBA倉庫在庫表']
        #! 初期化
        for row in range(5,2000,10):  # R to AAに対応
            ws.cells(row, 15).value = 0

        # 入力
        for row in filter_list:
            for i in range(5,5000,10):
                product_name = ws.cells(i,3).value
                if product_name == "" or product_name is None:
                    break
                else:
                    if product_name == row[0]:
                        value = int(ws.cells(i,15).value)
                        # 既に入力されているデータがある場合は、加算。
                        if value == 0:
                            ws.cells(i,15).value = int(row[1])
                        else:
                            ws.cells(i,15).value = value + int(row[1])

    def orders(self,orders_list):
        """ 塗屋への発注数を自動入力
        """
        ws = wb.sheets['発注・入庫']

        #! 初期化
        # 記入日時を初期化
        for row in range(3, 6):  # 3, 4, 5に対応
            for col in range(6, 20):  # R to AAに対応
                ws.cells(row, col).value = ""

        # 入庫数を初期化
        for row in range(7, 107):  # 7 to 106に対応
            for col in range(6, 20):  # R to AAに対応
                ws.cells(row, col).value = ""

        #! データの前準備
        #! ➀発注日リスト算出
        # 指定シリーズ以外除外
        filter_list = [row for row in orders_list if series_name in row]
        # 発注日順に並び替え
        sorted_list = sorted(filter_list, key=lambda x: datetime.datetime.strptime(x[0], "%Y/%m/%d"))
        # 日付のみ（重複除外）を算出
        day_list = list(set([row[0] for row in sorted_list]))
        # 念のために、日付を並び替え
        order_date_list = sorted(day_list, key=lambda x: datetime.datetime.strptime(x, "%Y/%m/%d")) 

        #! ➁各発注日の希望納期リストを算出
        preferred_delivery_date_list = [] # 希望納期リスト
        for day in order_date_list:
            # 発注日でフィルター
            filter2_list = [row for row in orders_list if row[0] == day]
            # 希望納期順に並び替え
            sorted2_list = sorted(filter2_list, key=lambda x: datetime.datetime.strptime(x[1], "%Y/%m/%d"))
            # 「希望納期」と「発注先」を取り出す（重複除外）
            day2_list = [row[1:3] for row in sorted2_list]
            day2_list = list(map(list,set(map(tuple,day2_list))))
            # 念のために発注日順に並び替え
            sorted_day2_list = sorted(day2_list, key=lambda x: datetime.datetime.strptime(x[0], "%Y/%m/%d"))
            preferred_delivery_date_list.append(sorted_day2_list)

        # 「発注日別」・「希望納期別」に入庫数を入力
        l = 0
        for i in range(len(order_date_list)):
            for j in range(len(preferred_delivery_date_list[i])):
                # 「発注日」、「希望納期」、「発注先」を記入
                ws.cells(3, 6 + l).value = order_date_list[i]
                ws.cells(4, 6 + l).value = preferred_delivery_date_list[i][j][0]
                ws.cells(5, 6 + l).value = preferred_delivery_date_list[i][j][1]

                # 日付フィルターを掛けて、必要なデータのみ準備
                filter3_list = [row for row in sorted_list if row[0] == order_date_list[i] and row[1] == preferred_delivery_date_list[i][j][0] and row[2] == preferred_delivery_date_list[i][j][1]]

                for row in filter3_list:
                    k = 7
                    while True: # シート探索
                        product_name = ws.cells(k,1).value
                        if product_name == "" or product_name is None: # データが無くなったら終了
                            break
                        elif product_name == row[3]:
                            value = ws.cells(k,6 + l).value
                            # 既に入力されているデータがある場合は、加算。
                            if value == "" or value is None:
                                ws.cells(k,6 + l).value = int(row[4])
                            else:
                                ws.cells(k,6 + l).value = int(ws.cells(k,6 + l).value) + int(row[4])
                        k += 1

                l += 1

    def receipts(self,receipts_list):
        """ 塗屋からの入庫数を自動入力
        """
        ws = wb.sheets['発注・入庫']

        #! 初期化
        # 記入日時を初期化
        for row in range(3, 6):  # 3, 4, 5に対応
            for col in range(23, 38):  # R to AAに対応
                ws.cells(row, col).value = ""

        # 入庫数を初期化
        for row in range(7, 107):  # 7 to 106に対応
            for col in range(23, 38):  # R to AAに対応
                ws.cells(row, col).value = ""

        #! データ前準備
        # filterによる指定シリーズ以外の行を削除
        filter_list = [row for row in receipts_list if series_name in row]
        # 出荷日順に並び替え
        sorted_list = sorted(filter_list, key=lambda x: datetime.datetime.strptime(x[1], "%Y/%m/%d"))
        # 日付のみを算出
        day_list = list(set([row[1] for row in sorted_list]))
        # 念のために、日付を並び替え
        sorted_day_list = sorted(day_list, key=lambda x: datetime.datetime.strptime(x, "%Y/%m/%d"))

        #! 入力
        # 日付を記入
        for i in range(len(sorted_day_list)):  # 3, 4, 5に対応
            ws.cells(3, 23 + i).value = sorted_day_list[i]
            ws.cells(5, 23 + i).value = "山家"

        # 日付別に入庫数を入力
        for i in range(len(sorted_day_list)):
            # 日付フィルターを掛けて、必要なデータのみ準備
            filter2_list = [row for row in sorted_list if sorted_day_list[i] == row[1]]
            for row in filter2_list:
                for j in range(7,1000): # シート探索
                    if ws.cells(j,1).value == "" or ws.cells(j,1).value == None: # データが無くなったら終了
                        break
                    else:
                        if ws.cells(j,1).value == row[3]:
                            # 既に入力されているデータがある場合は、加算。
                            if ws.cells(j,23 + i).value == "" or ws.cells(j,23 + i).value == None:
                                ws.cells(j,23 + i).value = int(row[4])
                            else:
                                ws.cells(j,23 + i).value = int(ws.cells(j,23 + i).value) + int(row[4])

    def deliveries(self,deliveries_list):
        """FBA納品表へFBA納品数の入力
        """
        ws = wb.sheets['FBA納品表']

        #! 初期化
        # 納品日・納品数を初期化
        for col in range(6, 21):  # F to Oに対応
            ws.cells(3, col).value = ""
            for row in range(5,1000,10):  # 5 to 995 step 10に対応
                ws.cells(row, col).value = ""

        #! データ前準備
        # filterによる指定シリーズ以外の行を削除
        filter_list = [row for row in deliveries_list if series_name in row]
        # 出荷日順に並び替え
        sorted_list = sorted(filter_list, key=lambda x: datetime.datetime.strptime(x[0], "%Y/%m/%d"))
        # 日付のみを算出
        day_list = list(set([row[0] for row in sorted_list]))
        # 念のために、日付を並び替え
        sorted_day_list = sorted(day_list, key=lambda x: datetime.datetime.strptime(x, "%Y/%m/%d"))

        #! 入力
        # 日付を記入
        for i in range(len(sorted_day_list)):
            ws.cells(3, 6 + i).value = sorted_day_list[i]

        # 日付別に入庫数を入力
        for i in range(len(sorted_day_list)):
            for row in sorted_list:
                if row[0] == sorted_day_list[i]:
                    for j in range(5,1000,10): # シート探索
                        if ws.cells(j,2).value == "" or ws.cells(j,2).value == None: # データが無くなったら終了
                            break
                        else:
                            # sku一致した場合、
                            if ws.cells(j,2).value == row[4]:
                                # 既に入力されているデータがある場合は、加算。
                                if ws.cells(j,6 + i).value == "" or ws.cells(j, 6+ i).value == None:
                                    ws.cells(j,6 + i).value = int(row[2])
                                else:
                                    ws.cells(j,6 + i).value = int(ws.cells(j,6 + i).value) + int(row[2])

    def orderQuantity(self,orderQuantity_list):
        """注文数自動算出の出力

        Args:
            List (_type_): [品名、シリーズ、色、現在庫、適正在庫数、発注時入数、注文残り]
        """
        for series in ["Natural","和食器"]:
            if series == "Natural":
                for color in ["ブラック","ブラウン","ベージュ","オフホワイト"]:
                    ws = wb.sheets[color]
                    n = 2
                    for row in orderQuantity_list:
                        if row[1] == series and row[2] == color:
                            orderCaseQuantity = 0
                            while True:
                                # 注文数 = 現在庫 + 注残 + 注文ケース数 * 箱入り数
                                orderQuantity = row[3] + row[6] + orderCaseQuantity * int(row[5])
                                if orderQuantity >= row[4]*1.5:
                                    ws.cells(n, 1).value = row[0] # 品名
                                    ws.cells(n, 2).value = row[1] # シリーズ
                                    ws.cells(n, 3).value = row[2] # 色
                                    ws.cells(n, 4).value = orderCaseQuantity # 発注ケース数
                                    ws.cells(n, 5).value = orderCaseQuantity*int(row[5]) # 個数
                                    ws.cells(n, 6).value = row[5] # 箱入り数
                                    ws.cells(n, 7).value = row[6] # 注残
                                    break
                                orderCaseQuantity += 1
                            n += 1
            elif series == "和食器":
                ws = wb.sheets[series]
                n = 2
                for row in orderQuantity_list:
                    if row[1] == series:
                        orderCaseQuantity = 0
                        while True:
                            # 注文数 = 現在庫 + 注残 + 注文ケース数 * 箱入り数
                            orderQuantity = row[3] + row[6] + orderCaseQuantity * int(row[5])
                            if orderQuantity >= row[4]*1.5:
                                ws.cells(n, 1).value = row[0] # 品名
                                ws.cells(n, 2).value = row[1] # シリーズ
                                ws.cells(n, 3).value = row[2] # 色
                                ws.cells(n, 4).value = orderCaseQuantity # 発注ケース数
                                ws.cells(n, 5).value = orderCaseQuantity*int(row[5]) # 個数
                                ws.cells(n, 6).value = row[5] # 箱入り数
                                ws.cells(n, 7).value = row[6] # 注残
                                break
                            orderCaseQuantity += 1
                        n += 1

class Add:
    def __init__(self):
        pass

    def inventory(self):
        """出庫数・在庫数をexcelファイルに更新
        """
        LIST = []

        ###! データを取得
        # 初期設定
        inventory_list = [] # 自社倉庫在庫数リスト
        FBA_inventory_list = [] # FBA倉庫在庫数リスト
        single_issue_list = [] # 単品出庫数リスト
        set_issue_list = [] # セット商品出庫数リスト
        FBA_set_issue_list = [] # FBAセット商品出庫数リスト

        print("更新用データ取得開始")
        window.refresh()
        num = 0
        for series_name in series_list:
            num += 1
            print(f"- {series_name}シリーズ ({num}/{len(series_list)})")
            window.refresh()
            file_name = f'data\\【新EC在庫表】{series_name}_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx'

            try:
                # 自社倉庫在庫数を算出
                df = pd.read_excel(file_name,sheet_name='自社倉庫在庫表')
                array = df.to_numpy()
                for i in range(len(array)):
                    if i > 1:
                        if str(array[i][0]) == "nan":
                            pass
                        else:
                            sku = array[i][1]
                            name = array[i][2]
                            inventory = array[i][5]
                            set_to_single_issue = array[i][7]
                            FBA_set_to_single_issue = array[i][19]
                            inventory_list.append([sku,name,inventory,set_to_single_issue,FBA_set_to_single_issue])

                # FBA倉庫在庫数を算出
                df = pd.read_excel(file_name,sheet_name='FBA倉庫在庫表')
                array = df.to_numpy()
                a = len(array)
                for i in range(len(array)):
                    if i > 1:
                        if str(array[i][1]) == "nan":
                            pass
                        else:
                            sku = array[i][1]
                            name = array[i][2]
                            inventory = array[i][5]
                            FBA_inventory_list.append([sku,name,inventory])

                # 単品出庫数を算出
                df = pd.read_excel(file_name,sheet_name='単品出庫表')
                array = df.to_numpy()
                for i in range(len(array)):
                    if i > 1:
                        if str(array[i][1]) == "nan":
                            pass
                        else:
                            sku = array[i][1]
                            name = array[i][2]
                            issue = array[i][6]
                            single_issue_list.append([sku,name,issue])

                # セット商品出庫数
                df = pd.read_excel(file_name,sheet_name='セット商品出庫表')
                array = df.to_numpy()
                for i in range(len(array)):
                    if i > 1:
                        if str(array[i][1]) == "nan":
                            pass
                        else:
                            sku = array[i][1]
                            name = array[i][2]
                            issue = array[i][6]
                            set_issue_list.append([sku,name,issue])

                # セット商品FBA出庫数
                df = pd.read_excel(file_name,sheet_name='セット商品FBA出庫表')
                array = df.to_numpy()
                for i in range(len(array)):
                    if i > 1:
                        if str(array[i][1]) == "nan":
                            pass
                        else:
                            sku = array[i][1]
                            name = array[i][2]
                            issue = array[i][6]
                            FBA_set_issue_list.append([sku,name,issue])

            except FileNotFoundError:
                    print(f"File not found: {file_name}")
                    pass

        print("既存データ取得開始")
        window.refresh()
        try:
            LIST = setlist_excel("conf/在庫数出庫数の履歴.xlsx",sheet_name="Sheet",header_skip=False)
        except KeyError:
            LIST = setlist_excel("conf/在庫数出庫数の履歴.xlsx",sheet_name="Sheet1",header_skip=False)


        print("ヘッダー追加")
        window.refresh()
        # 「自社倉庫現在庫：指定月」ヘッダー追加
        if not f"自社倉庫在庫数：{selected_date.year}年{selected_date.month}月" in LIST[0]:
            LIST[0].append(f"自社倉庫在庫数：{selected_date.year}年{selected_date.month}月")

        # 「FBA倉庫現在庫：指定月」ヘッダー追加
        if not f"FBA倉庫在庫数：{selected_date.year}年{selected_date.month}月" in LIST[0]:
            LIST[0].append(f"FBA倉庫在庫数：{selected_date.year}年{selected_date.month}月")

        # 「出庫数：指定月」ヘッダー追加
        if not f"出庫数：{selected_date.year}年{selected_date.month}月" in LIST[0]:
            LIST[0].append(f"出庫数：{selected_date.year}年{selected_date.month}月")

        # 「セット商品を単品に分解した出庫数：指定月」ヘッダー追加
        if not f"セット商品を単品に分解した出庫数：{selected_date.year}年{selected_date.month}月" in LIST[0]:
            LIST[0].append(f"セット商品を単品に分解した出庫数：{selected_date.year}年{selected_date.month}月")

        # 列長を統一する
        max_col = len(LIST[0]) # 最大列数
        for i in range(len(LIST)):
            now_col = len(LIST[i])
            if now_col < max_col:
                for j in range(max_col - now_col):
                    LIST[i].append("")

        print("入力データ設定")
        window.refresh()
        for i in range(len(LIST[0])):
            # 「自社倉庫在庫数：指定月」入力
            if LIST[0][i] == f"自社倉庫在庫数：{selected_date.year}年{selected_date.month}月":
                for j in range(len(LIST)):
                    for k in range(len(inventory_list)):
                        # 製品名とSKUが一致
                        if LIST[j][1] == inventory_list[k][1] and LIST[j][5] == inventory_list[k][0]:
                            LIST[j][i] = inventory_list[k][2]
            # 「FBA倉庫現在庫：指定月」入力
            if LIST[0][i] == f"FBA倉庫在庫数：{selected_date.year}年{selected_date.month}月":
                for j in range(len(LIST)):
                    for k in range(len(FBA_inventory_list)):
                        # 製品名とSKUが一致
                        if LIST[j][1] == FBA_inventory_list[k][1] and LIST[j][5] == FBA_inventory_list[k][0]:
                            LIST[j][i] = FBA_inventory_list[k][2]
            # 「出庫数：指定月」入力
            if LIST[0][i] == f"出庫数：{selected_date.year}年{selected_date.month}月":
                for j in range(len(LIST)):
                    # 単品出庫数を入力
                    if LIST[j][3] == "単品":
                        for k in range(len(single_issue_list)):
                            # SKUが一致
                            if LIST[j][5] == single_issue_list[k][0]:
                                LIST[j][i] = single_issue_list[k][2]
                    # セット商品出庫数・FBAセット商品数を入力
                    elif LIST[j][3] == "セット":
                        # セット商品出庫数を入力
                        for k in range(len(set_issue_list)):
                            # SKUが一致
                            if LIST[j][5] == set_issue_list[k][0]:
                                if LIST[j][i] is None or LIST[j][i] == "":
                                    LIST[j][i] = int(set_issue_list[k][2])
                                else:
                                    LIST[j][i] = int(LIST[j][i]) + int(set_issue_list[k][2])
                        # FBAセット商品出庫数を入力
                        for k in range(len(FBA_set_issue_list)):
                            # SKUが一致
                            if LIST[j][5] == FBA_set_issue_list[k][0]:
                                if LIST[j][i] is None or LIST[j][i] == "":
                                    LIST[j][i] = int(FBA_set_issue_list[k][2])
                                    pass
                                else:
                                    LIST[j][i] = int(LIST[j][i]) + int(FBA_set_issue_list[k][2])
            # 「セット商品を単品に分解した出庫数：指定月」入力
            if LIST[0][i] == f"セット商品を単品に分解した出庫数：{selected_date.year}年{selected_date.month}月":
                for j in range(len(LIST)):
                    # セット商品を単品に分解した出庫数を入力
                    for k in range(len(inventory_list)):
                        # 製品名とSKUが一致
                        if LIST[j][1] == inventory_list[k][1] and LIST[j][5] == inventory_list[k][0]:
                            LIST[j][i] = int(inventory_list[k][3]) + int(inventory_list[k][4])

        #! xlsxに書き込み
        print("書き込み開始")
        window.refresh()
        save_excel("conf/在庫数出庫数の履歴.xlsx",LIST)

        #! googledriveへアップロード
        print("アップロード開始")
        window.refresh()
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                        search_folder_name_list=["conf"],
                        upload_folder_name="conf",
                        upload_file_name_list=["在庫数出庫数の履歴.xlsx"]
                        )

        log(["在庫数出庫数の履歴.xlsx 更新完了"])
        window.refresh()

def copySheet(open_filepath,save_filepath,sheet_number):
    print("開始")
    window.refresh()

    # Excelをバックグラウンドで開く
    app = xw.App(visible=False)  # Excelの可視性をFalseに設定

    # 新規Excelファイル作成
    wb = xw.Book()
    wb.save(save_filepath)
    # デフォルトシート名取得
    sheet_names = [sheet.name for sheet in wb.sheets]
    wb.close()

    # コピー元ファイル
    wb_a = app.books.open(open_filepath)
    # コピー先ファイル
    wb_z = app.books.open(save_filepath)

    # シートコピー
    wb_a.sheets[sheet_number].copy(before=wb_z.sheets[0])

    # シート名変更
    #todo --- 2024/9/10 変更-----------------
    try:
        ws = wb_z.sheets['発注・入庫']
        ws.name = '注残'
    except Exception as e:
        print(f"シート名変更エラー:{e}")
        print("再度変更します。")
        ws = wb_z.sheets[0]
        ws.name = '注残'
    #todo end --------------------------------

    # デフォルトシート削除
    print("デフォルトシート削除")
    window.refresh()
    for sheet_name in sheet_names:
        try:
            wb_z.sheets[sheet_name].delete()
        except Exception as e:
            print(e)


    # コピー元のシートを選択
    source_sheet = wb_a.sheets['発注・入庫']

    # 特定の範囲 "A1:AI200" の値を取得
    values_to_copy = source_sheet.range('A1:AI200').value

    # 新しいシートを作成（または既存のシートを指定）
    target_sheet = wb_z.sheets['注残']

    # 値を新しいシートにペースト
    target_sheet.range('A1').value = values_to_copy

    # コピー先ファイルを保存
    wb_z.save()

    # ファイルを閉じる
    wb_a.close()
    wb_z.close()

    app.quit()
    print("完了")
    window.refresh()

def log(list:list,time_switch=True):
    """ログ作成

    Args:
        product_name (str): 対象成形機番号
        e (str): 内容
    """
    y=datetime.datetime.now().year
    m=datetime.datetime.now().month
    d=datetime.datetime.now().day
    hour=datetime.datetime.now().hour
    min=datetime.datetime.now().minute
    sec=datetime.datetime.now().second

    log_filename = "log_{}-{}-{}.log".format(str(y).zfill(4),str(m).zfill(2),str(d).zfill(2))
    mode_="w" if log_filename in glob.glob("logs/*") else "a"

    # 出力する内容
    time_now = str(hour).zfill(2)+":"+str(min).zfill(2)+":"+str(sec).zfill(2)
    body=""

    # 出力内容を一つにまとめる
    for i in range(len(list)):
        body = body + str(list[i])

    # ログファイルに書き込み
    with open("logs/"+log_filename,mode=mode_,newline="\n") as logfile:
        if time_switch:
            logfile.writelines(["{} {}\n".format(time_now,body)])
        else:
            logfile.writelines(["{}\n".format(body)])

    print(body)

###! GUI画面作成
sg.theme('BlueMono')

date_layout = [[sg.Text("日付指定",font=('Helvetica', 12, 'bold')),
                sg.Combo(values=[i for i in range(2022,2100)],default_value=today.year,size=(5,1),key='selectdate_year',readonly=True),
                sg.Text("年"),
                sg.Combo(values=[i for i in range(1,13)],default_value=today.month,size=(5,1),key='selectdate_month',readonly=True,enable_events=True),
                sg.Text("月"),
                sg.Combo(values=[i for i in range(1,32)],default_value=today.day,size=(5,1),key='selectdate_day',readonly=True),
                sg.Text("日")
                ]]

# まとめレイアウト
layout = [
    [sg.MenuBar([['ファイル',['保存先フォルダを開く']],
                ['機能',['在庫表自動入力','新規在庫表作成',"フォーマット更新",'注文数自動算出']],
                ['ヘルプ',['操作手順','注意事項','概要（リンクを開く）','製品情報']]])],
    date_layout,
    [sg.Frame('事前準備',[[sg.Button('Amazon Seller Centralを開く',size=(30,1),key="open_amazonHP_btn")],
                            [sg.Input("Amazonデータファイルを選択してください。",key="input_file"),sg.FileBrowse('選択')],])],
    [sg.Frame("在庫表自動入力",[[sg.Button("一月分",size=(15,1),key="run_1month_btn")],[sg.Text('※指定月以外の日は除外されます。')]])],
    [sg.Frame("追加機能",[[sg.Button("新規在庫表作成",size=(20,2),key="mkfile_btn"),
                            sg.Button("フォーマット更新",size=(20,2),key="update_format_btn")],
                        [sg.Button("注文数自動算出\n(対象:Natural,和食器)",size=(20,2),key="order_quantity_calculation_btn")]])],
    [sg.Text("出力",font=('Helvetica', 12, 'bold'))],
    [sg.Output(size=(70,12))]
    ]

window = sg.Window(f'{PROGRAM_NAME} {PROGRAM_VERSION}',layout=layout,size=(600,600),finalize=True)

###! メインプログラム
log([f"{PROGRAM_NAME} を起動。"])

while True:
    # イベント取得
    event, values = window.read()
    # ウィンドウクローズか、Exitボタン押下時
    if event is None:
        break

    if event in ["selectdate_year","selectdate_month"]:
        last_day = calendar.monthrange(int(values["selectdate_year"]),int(values["selectdate_month"]))[1]
        day_list = [i for i in range(1,last_day + 1)]
        if int(values["selectdate_year"]) <= datetime.datetime.now().year and int(values["selectdate_month"]) < datetime.datetime.now().month:
            window.find_element("selectdate_day").Update(values=day_list,value=last_day)
        elif int(values["selectdate_year"]) == datetime.datetime.now().year and int(values["selectdate_month"]) == datetime.datetime.now().month:
            window.find_element("selectdate_day").Update(values=day_list,value=datetime.datetime.now().day)
        else:
            window.find_element("selectdate_day").Update(values=day_list,value=1)

    if event == "open_amazonHP_btn":
        url = "https://sellercentral-japan.amazon.com/ap/signin?openid.return_to=https%3A%2F%2Fsellercentral-japan.amazon.com%2Forder-reports-and-feeds%2Freports%2Fref%3Dxx_orderrpt_dnav_xx&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=sc_jp_amazon_com_v2&openid.mode=checkid_setup&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&mons_redirect=sign_in&ssoResponse=eyJ6aXAiOiJERUYiLCJlbmMiOiJBMjU2R0NNIiwiYWxnIjoiQTI1NktXIn0.I25m_rMseU3vR0Reu8qBp29d1tR7CHvyZWqr0PmsXdfJOncoXxe-sw.J1RMA2GPZBmqN2Jx.4K60fb1A9lxDac72uZJfP5Ryj9IR7QSX8hPZR4t94y-7i2DSLOHRXQIIndsaOspSpYCLcs_e3t7GofUNofDlDHNxTJ66R20O99s6GB83-lJpFa3nDTTTdjMF_1O1vdBOOXp-DfVatbYehMhgeb-ftv4w81-t9-v-woOac2LhQ0f1FhX0d3dByjVJCtBw0kT3DTGQvEET.OAHMK-PZBfvsz47JK7ZYMQ"  # 開きたいウェブサイトのURLを指定してください
        webbrowser.open(url)

    if "mkfile" in event or "新規在庫表作成" in event:
        # アプリを終了
        close_window(["chrome","excel"])
        # 互換性確認
        version_matching_confirmation()

        log(["-----"],time_switch=False)
        log(["新規在庫表作成開始"])
        selected_date = datetime.datetime(values['selectdate_year'],values['selectdate_month'],values['selectdate_day'])
        # 先月
        one_month_ago = datetime.datetime(values['selectdate_year'],values['selectdate_month'],15) - relativedelta(months=1)
        remove_file(["在庫表"],"data")
        #! 最終確認
        choice = sg.popup_ok_cancel(f'新規の在庫表を作成します。（{selected_date.year}年{selected_date.month}月分)\n'
                            '※既にGoogleDriveに在庫表がある場合は、新規在庫表に置き換えられます。\n'
                            '  問題が無ければ、"OK"を押してください。新規在庫表の作成を開始します。\n'
                            '  中止する場合は、"NG"を押してください。アプリを終了します。')
        # Check the user's selection and act accordingly
        if choice == 'OK':
            pass
        else:
            log(["新規在庫表作成を中断"])
            log(["ECAIISを終了します"])
            # Terminate the program
            sg.popup('ECAIISを終了します。')
            exit()

        #! ダウンロード
        # 共通フォーマット
        download_for_googledrive(saving_foldername="conf",search_list=["【新EC在庫表】共通フォーマット.xlsx"],folder_id="177__7xnQYhxmjyfs6VGnEobkLI1gLI4g")
        # skulist
        download_for_googledrive(saving_foldername="conf",search_list=["商品情報.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")
        # sales_inventory
        download_for_googledrive(saving_foldername="conf",search_list=["在庫数出庫数の履歴.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")
        # 前月の注残数
        download_for_googledrive(saving_foldername="temp",search_list=[f"（自動更新）注残数_{str(one_month_ago.year)}-{str(one_month_ago.month).zfill(2)}.xlsx"],folder_id="1uQULhUO5-LUGrAg9Ihy_KRCkkwHpR0We")

        #! データ格納
        # 商品詳細リスト
        window.refresh()
        product_details_list = setlist_excel("conf/商品情報.xlsx",sheet_name="Sheet1",header_skip=True)

        # 商品の売上と過去出庫数
        window.refresh()
        sales_inventory_list = setlist_excel("conf/在庫数出庫数の履歴.xlsx",sheet_name="Sheet",header_skip=False)

        # 前月の注残数
        PP = Preparation()
        outstanding_orders_list = PP.lastMonthOrderRest()

        series_list = list(set([row[2] for row in product_details_list]))

        print(series_list)
        window.refresh()
        #! シリーズ毎に在庫表を作成し、必要なデータを入力する
        num = 0
        for series_name in series_list:
            num += 1
            print(f"--- {series_name}在庫表作成 ({num}/{len(series_list)}) ---")
            window.refresh()
            # excelをアプリで開く
            app = xw.App(visible=False)
            # 計算を手動に設定
            app.calculation = 'manual'

            # フォーマットから複製
            print("フォーマットから在庫表を作成")
            window.refresh()
            source = f'conf\\【新EC在庫表】共通フォーマット.xlsx'
            file_name = f'data\\【新EC在庫表】{series_name}_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx'
            shutil.copy(source,file_name)

            # 開く
            wb = app.books.open(f'{file_name}',read_only=False)
            time.sleep(5)

            print("各シートデータ更新")
            window.refresh()
            #! 「単品出庫表」シートの日付（初日）を変更。（他のシートの日付はこのセルに付随する）
            ws = wb.sheets['単品出庫表']
            ws.cells(2,8).value =  f"{selected_date.year}/{selected_date.month}/1" # 日付

            #! 「単品商品リスト」シートに追加
            ws = wb.sheets['単品商品リスト']
            row = 2 # 単品リスト入力位置（行）
            for i in range(len(product_details_list)):
                # シリーズ名が同じ
                if product_details_list[i][2] == series_name:
                    if product_details_list[i][3] == "単品":
                        ws.cells(row,1).value = product_details_list[i][1] # 品名
                        ws.cells(row,2).value = product_details_list[i][0] # 品番
                        ws.cells(row,3).value = product_details_list[i][4] # 色
                        ws.cells(row,4).value = product_details_list[i][5] # sku
                        ws.cells(row,5).value = product_details_list[i][10] # 発注時入数
                        ws.cells(row,6).value = product_details_list[i][11] # 発注時単価
                        # 列探索
                        for j in range(len(sales_inventory_list[0])):
                            # 該当データ列あり、
                            if sales_inventory_list[0][j] == f"自社倉庫在庫数：{one_month_ago.year}年{one_month_ago.month}月":
                                for k in range(len(sales_inventory_list)):
                                    # 商品番号と品名が「単品商品リストsheet」に一致した場合、
                                    if sales_inventory_list[k][0] == product_details_list[i][0] \
                                        and sales_inventory_list[k][1] == product_details_list[i][1]:
                                        ws.cells(row,7).value = sales_inventory_list[k][j] # 先月在庫数
                                        break

                        # 次の入力位置へ（一つ行を下に）
                        row += 1

            #! 「セット商品リスト」シートに追加
            ws = wb.sheets['セット商品リスト']
            row = 2
            for i in range(len(product_details_list)):
                # シリーズ名が同じ
                if product_details_list[i][2] == series_name:
                    if product_details_list[i][3] == "セット":
                        ws.cells(row,1).value = product_details_list[i][1] # セット商品名
                        ws.cells(row,2).value = product_details_list[i][0] # 品番
                        ws.cells(row,3).value = product_details_list[i][4] # 色
                        ws.cells(row,4).value = product_details_list[i][5] # sku
                        for j in range(0,10):
                            ws.cells(row, 2*j+5).value = product_details_list[i][2*j+12] # 品名1
                            ws.cells(row, 2*j+6).value = product_details_list[i][2*j+13] # 数量1
                        # 列探索
                        for j in range(len(sales_inventory_list[0])):
                            # 該当データ列あり、
                            if sales_inventory_list[0][j] == f"FBA倉庫在庫数：{one_month_ago.year}年{one_month_ago.month}月":
                                for k in range(len(sales_inventory_list)):
                                    # 商品番号と品名が「単品商品リストsheet」に一致した場合、
                                    if sales_inventory_list[k][0] == product_details_list[i][0] \
                                        and sales_inventory_list[k][1] == product_details_list[i][1]:
                                        ws.cells(row,25).value = sales_inventory_list[k][j] # 先月在庫数

                        row += 1

            #! 「販売予測」シートに追加
            ws = wb.sheets['販売予測']
            row = 4
            one_month_ago = selected_date - relativedelta(months=1)
            two_month_ago = selected_date - relativedelta(months=2)
            three_month_ago = selected_date - relativedelta(months=3)
            for i in range(len(product_details_list)):
                # シリーズ名が同じ
                if product_details_list[i][2] == series_name:
                    if product_details_list[i][3] == "単品":
                        for j in range(len(sales_inventory_list)):
                            if sales_inventory_list[j][0] == product_details_list[i][0]:
                                for k in range(len(sales_inventory_list[0])):
                                    if f"出庫数：{one_month_ago.year}年{one_month_ago.month}月" == sales_inventory_list[0][k]:
                                        issue = 0
                                        issue_0 = ws.cells(row,7).value
                                        issue_1 = sales_inventory_list[j][k]
                                        issue = issue if issue_0 is None or issue_0 == "" or issue_0 == 0 else issue + int(issue_0)
                                        issue = issue if issue_1 is None or issue_1 == "" or issue_1 == 0 else issue + int(issue_1)
                                        ws.cells(row,7).value = issue
                                    if f"出庫数：{two_month_ago.year}年{two_month_ago.month}月" == sales_inventory_list[0][k]:
                                        issue = 0
                                        issue_0 = ws.cells(row,6).value
                                        issue_1 = sales_inventory_list[j][k]
                                        issue = issue if issue_0 is None or issue_0 == "" or issue_0 == 0 else issue + int(issue_0)
                                        issue = issue if issue_1 is None or issue_1 == "" or issue_1 == 0 else issue + int(issue_1)
                                        ws.cells(row,6).value = issue
                                    if f"出庫数：{three_month_ago.year}年{three_month_ago.month}月" == sales_inventory_list[0][k]:
                                        issue = 0
                                        issue_0 = ws.cells(row,5).value
                                        issue_1 = sales_inventory_list[j][k]
                                        issue = issue if issue_0 is None or issue_0 == "" or issue_0 == 0 else issue + int(issue_0)
                                        issue = issue if issue_1 is None or issue_1 == "" or issue_1 == 0 else issue + int(issue_1)
                                        ws.cells(row,5).value = issue
                                    if f"セット商品を単品に分解した出庫数：{one_month_ago.year}年{one_month_ago.month}月" == sales_inventory_list[0][k]:
                                        issue = 0
                                        issue_0 = ws.cells(row,7).value
                                        issue_1 = sales_inventory_list[j][k]
                                        issue = issue if issue_0 is None or issue_0 == "" or issue_0 == 0 else issue + int(issue_0)
                                        issue = issue if issue_1 is None or issue_1 == "" or issue_1 == 0 else issue + int(issue_1)
                                        ws.cells(row,7).value = issue
                                    if f"セット商品を単品に分解した出庫数：{two_month_ago.year}年{two_month_ago.month}月" == sales_inventory_list[0][k]:
                                        issue = 0
                                        issue_0 = ws.cells(row,6).value
                                        issue_1 = sales_inventory_list[j][k]
                                        issue = issue if issue_0 is None or issue_0 == "" or issue_0 == 0 else issue + int(issue_0)
                                        issue = issue if issue_1 is None or issue_1 == "" or issue_1 == 0 else issue + int(issue_1)
                                        ws.cells(row,6).value = issue
                                    if f"セット商品を単品に分解した出庫数：{three_month_ago.year}年{three_month_ago.month}月" == sales_inventory_list[0][k]:
                                        issue = 0
                                        issue_0 = ws.cells(row,5).value
                                        issue_1 = sales_inventory_list[j][k]
                                        issue = issue if issue_0 is None or issue_0 == "" or issue_0 == 0 else issue + int(issue_0)
                                        issue = issue if issue_1 is None or issue_1 == "" or issue_1 == 0 else issue + int(issue_1)
                                        ws.cells(row,5).value = issue
                                break
                        row += 1

            # 「発注・入庫シート」に「注残」入力するために一時保存
            try:
                log(["注残数入力のために一度在庫表を保存"])
                window.refresh()

                # 計算を自動に戻して、全シートの再計算を実行
                app.calculation = 'automatic'
                for sheet in wb.sheets:
                    sheet.api.Calculate()  # 各シートの再計算

                #ファイルの保存とブックのクローズ
                wb.save(file_name)
                wb.close()
                app.quit()
                time.sleep(5)
            except Exception as e:
                app.quit()
                print("エラー：",e)
                window.refresh()
                pass

            # excelをアプリで開く
            app = xw.App(visible=False)
            # 計算を手動に設定
            app.calculation = 'manual'

            #! 「発注・入庫シート」に「注残」入力
            print("注残数を入力")
            window.refresh()
            # 開く
            wb = app.books.open(f'{file_name}',read_only=False)
            time.sleep(5)

            ws = wb.sheets['発注・入庫']
            for row in outstanding_orders_list:
                if row[1] == series_name:
                    for j in range(7,1000): # シート探索
                        a = ws.cells(j,1).value
                        if ws.cells(j,1).value == "" or ws.cells(j,1).value == None: # データが無くなったら終了
                            break
                        else:
                            if ws.cells(j,1).value == row[0]:
                                # 既に入力されているデータがある場合は、加算。
                                if ws.cells(j,5).value == "" or ws.cells(j,5).value == None:
                                    ws.cells(j,5).value = int(row[3])
                                else:
                                    ws.cells(j,5).value = int(ws.cells(j,5).value) + int(row[3])

            try:
                log(["在庫表保存"])
                window.refresh()

                # 計算を自動に戻して、全シートの再計算を実行
                app.calculation = 'automatic'
                for sheet in wb.sheets:
                    sheet.api.Calculate()  # 各シートの再計算

                #ファイルの保存とブックのクローズ
                wb.save(file_name)
                wb.close()
                app.quit()
                time.sleep(5)

                log(["保存完了"])
                window.refresh()

            except Exception as e:
                app.quit()
                print("エラー：",e)
                window.refresh()
                pass

        log(["全シリーズの新規在庫表作成完了"])

        log(["GoogleDriveへのアップロード開始"])
        window.refresh()
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["ECAIIS",f"{selected_date.year}年{selected_date.month}月在庫表"],
                            upload_folder_name="data",
                            upload_file_name_list=["在庫表",f"{selected_date.year}年{str(selected_date.month).zfill(2)}月"]
                            )
        log(["アップロード完了"])
        window.refresh()

    #! フォーマットの更新を行なう
    # 更新対象フォーマット：「返品_フォーマット」、「サンプル・不良・調整_フォーマット」、「FBA納品_フォーマット」、「発注_フォーマット」、「山家様_納品CSV作成フォーマット.xlsx」
    if "update_format" in event or "フォーマット更新" in event:
        # アプリを終了
        close_window(["excel"])
        # ファイルの削除
        remove_file(["フォーマット",".xlsx"],"conf")

        log(["各フォーマットの更新開始"])
        window.refresh()

        log(["GoogleDriveから「各フォーマット」と「商品情報.xlsx」、「各フォーマットの表示商品リスト.xlsx」をダウンロード"])
        window.refresh()
        #! 1.GoogleDriveから「各フォーマット」と「商品情報.xlsx」、「各フォーマットの表示商品リスト.xlsx」をダウンロード
        download_for_googledrive(saving_foldername="conf",search_list=["返品_フォーマット.xlsx"],folder_id="1WszXFq7e-1yKvQ8gs45-09oWQYZZLeBM")
        download_for_googledrive(saving_foldername="conf",search_list=["サンプル・不良・調整_フォーマット.xlsx"],folder_id="1RST89cKOd3U91EKOEvLATElfL5bQ7bqK")
        download_for_googledrive(saving_foldername="conf",search_list=["FBA納品_フォーマット.xlsx"],folder_id="1Nz2Gg6LABJp4pxiXY93VRS5NFSh54xVU")
        download_for_googledrive(saving_foldername="conf",search_list=["発注_フォーマット.xlsx"],folder_id="1uQULhUO5-LUGrAg9Ihy_KRCkkwHpR0We")
        download_for_googledrive(saving_foldername="conf",search_list=["山家様_納品CSV作成フォーマット.xlsx"],folder_id="1-x9H9Ixt7tAKUnmQ0F3qilEHfVkuMWxU")
        download_for_googledrive(saving_foldername="conf",search_list=["商品情報.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")
        download_for_googledrive(saving_foldername="conf",search_list=["各フォーマットの表示商品リスト.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")

        # skulistを格納
        skulist = setlist_excel("conf/商品情報.xlsx",sheet_name="Sheet1",header_skip=True)
        # selectedproduct_in_formatを格納
        selectedproduct_in_format = setlist_excel("conf/各フォーマットの表示商品リスト.xlsx",sheet_name="Sheet1",header_skip=True)

        #! 2.フォーマットを開き、商品リストの変更を行なう
        format_names = ["返品_フォーマット.xlsx","サンプル・不良・調整_フォーマット.xlsx","FBA納品_フォーマット.xlsx","発注_フォーマット.xlsx","山家様_納品CSV作成フォーマット.xlsx"]
        n = 6
        for format_name in format_names:
            log([f"{format_name}を更新開始"])
            window.refresh()
            #! 商品リストの準備
            list1 = [row for row in selectedproduct_in_format if int(row[n]) == 1] # 返品
            #! フォーマット準備
            file_name = f'conf\\{format_name}'
            # excelをアプリで開く
            app = xw.App(visible=False)
            wb = app.books.open(f'{file_name}',read_only=False)
            ws = wb.sheets['商品リスト']

            # 初期化
            ws.range("A2:E1000").value = ""

            # 最新版skuリスト入力
            # 入力
            for i in range(len(list1)):
                ws.range(f"A{i+2}").value = [list1[i][j] for j in range(1,6)]

            # 計算を自動に戻して、全シートの再計算を実行
            app.calculation = 'automatic'
            for sheet in wb.sheets:
                sheet.api.Calculate()  # 各シートの再計算

            #ファイルの保存とブックのクローズ
            wb.save(file_name)
            wb.close()
            log([f'ファイルの保存が完了しました。ファイル名「{file_name}」'])
            window.refresh()
            app.quit()
            time.sleep(5)

            n += 1

        #! 3.更新した各フォーマットをGoogleDriveへアップロード
        log(["GoogleDriveへアップロード開始"])
        window.refresh()
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["返品数データ"],
                            upload_folder_name="conf",
                            upload_file_name_list=["返品_フォーマット.xlsx"]
                            )
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["サンプル・不良・調整データ"],
                            upload_folder_name="conf",
                            upload_file_name_list=["サンプル・不良・調整_フォーマット.xlsx"]
                            )
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["FBA倉庫納品数データ"],
                            upload_folder_name="conf",
                            upload_file_name_list=["FBA納品_フォーマット.xlsx"]
                            )
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["塗屋発注数データ"],
                            upload_folder_name="conf",
                            upload_file_name_list=["発注_フォーマット.xlsx"]
                            )
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["塗屋入庫数データ"],
                            upload_folder_name="conf",
                            upload_file_name_list=["山家様_納品CSV作成フォーマット.xlsx"]
                            )
        log(["アップロード完了"])
        window.refresh()

    if "order_quantity_calculation" in event or "注文数自動算出" in event:
        log(["注文数自動算出を開始します。"])
        targetSeries = ["Natural","和食器"]

        selected_date = datetime.datetime(values['selectdate_year'],values['selectdate_month'],values['selectdate_day'])

        remove_file([".xlsx"],"data")

        #! データファイルダウンロード
        log([f"データファイルダウンロード開始"])
        window.refresh()
        log([f"1.最新版在庫表ダウンロード開始"])
        for series_name in targetSeries:
            window.refresh()
            download_for_googledrive(saving_foldername="data",search_list=["【新EC在庫表】",series_name,".xlsx",f"{selected_date.year}年{str(selected_date.month).zfill(2)}月"])
        log([f"2.注文数自動算出の出力フォーマットダウンロード開始"])
        window.refresh()
        download_for_googledrive(saving_foldername="conf",search_list=["注文数自動算出","フォーマット",".xlsx"])

        print("3.商品情報.xlsx")
        window.refresh()
        download_for_googledrive(saving_foldername="conf",search_list=["商品情報.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")

        # フォーマットから複製
        source = f'conf\\注文数自動算出_フォーマット.xlsx'
        file_name = f'data\\注文数自動算出_{datetime.datetime.now().year}年{str(datetime.datetime.now().month).zfill(2)}月{str(datetime.datetime.now().day).zfill(2)}日.xlsx'
        shutil.copy(source,file_name)

        orderQuantity = []
        num = 0
        for series_name in targetSeries:
            num += 1
            print(f"--- {series_name}シリーズ ({num}/{len(targetSeries)}) ---")
            window.refresh()
            try:
                P = Preparation()
                orderQuantity = P.orderQuantity(orderQuantity)
            except Exception as e:
                print(e)

        # excelをアプリで開く
        app = xw.App(visible=False)
        wb = app.books.open(f'{file_name}',read_only=False)
        # 計算を手動に設定
        app.calculation = 'manual'
        # 発注数を出力
        I = Input()
        I.orderQuantity(orderQuantity)

        try:
            # 計算を自動に戻して、全シートの再計算を実行
            app.calculation = 'automatic'
            for sheet in wb.sheets:
                sheet.api.Calculate()  # 各シートの再計算

            #ファイルの保存とブックのクローズ
            wb.save(file_name)
            wb.close()
            log(["ファイル保存完了"])
            window.refresh()
            app.quit()
            time.sleep(5)
        except Exception as e:
            app.quit()
            print("エラー：",e)
            window.refresh()
            pass

        #! 6. GoogleDriveへのアップロード
        # 在庫表をアップロード
        log(["GoogleDriveへのアップロード開始"])
        window.refresh()
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["注文数自動算出"],
                            upload_folder_name="data",
                            upload_file_name_list=["注文数自動算出",str(datetime.datetime.now().year),str(datetime.datetime.now().month),str(datetime.datetime.now().day)]
                            )
        log(["アップロード完了"])
        window.refresh()
        log(["注文数自動算出完了"])
        window.refresh()

    if "run_1month" in event or '在庫表自動入力' in event:
        #! 0. 初期動作
        # 指定フォルダ内のファイルを削除
        log(["昔のファイルを削除します。"])
        window.refresh()
        remove_file([".xlsx"],"data")
        remove_file([".csv",".xlsx"],"temp")

        # アプリを終了
        close_window(["chrome","excel"])
        # 互換性確認
        version_matching_confirmation()

        #! 0-1.日付設定
        selected_date = datetime.datetime(values['selectdate_year'],values['selectdate_month'],values['selectdate_day'])

        log(["１か月分実行開始"])
        window.refresh()
        # その月の初日
        start_of_month = selected_date.replace(day=1)
        # その月の最終日
        end_of_month = selected_date.replace(day=calendar.monthrange(selected_date.year, selected_date.month)[1])
        # 月初から月末までの日付をリストにする
        dates_list = [start_of_month + datetime.timedelta(days=i) for i in range((end_of_month - start_of_month).days + 1)]

        delete_list = []
        # 選択日と違う月の場合の日付リストから削除
        for date in dates_list:
            if not date.year == selected_date.year or not date.month == selected_date.month:
                delete_list.append(date)
            elif date.year == selected_date.year and date.month == selected_date.month:
                if date.day > selected_date.day:
                    delete_list.append(date)

        for delete_date in delete_list:
            dates_list.remove(delete_date)

        # 日付のフォーマット
        formatted_dates = [date.strftime('%Y-%m-%d') for date in dates_list]
        log([f"対象日リスト：{formatted_dates}"])
        window.refresh()

        #! 0-2.最新版の資料をダウンロード
        log([f"最新版在庫表、skulist、注残更新用フォーマットをダウンロード開始"])
        window.refresh()
        print("1. 在庫表")
        window.refresh()
        download_for_googledrive(saving_foldername="data",search_list=["【新EC在庫表】",".xlsx",f"{date.year}年{str(date.month).zfill(2)}月"])
        print("2. 商品情報.xlsx")
        window.refresh()
        download_for_googledrive(saving_foldername="conf",search_list=["商品情報.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")
        print("3. 在庫数出庫数の履歴.xlsx")
        window.refresh()
        download_for_googledrive(saving_foldername="conf",search_list=["在庫数出庫数の履歴.xlsx"],folder_id="1NkaUWEm69qoMglVUajLmUDSnPMQbPdWI")

        log([f"skulistをリストに格納"])
        window.refresh()
        skulist = setlist_excel("conf/商品情報.xlsx",sheet_name="Sheet1",header_skip=True)
        series_list = list(set([row[2] for row in skulist]))

        log([f"ログイン情報を取得"])
        window.refresh()
        # ログイン情報取得
        login_data = decryption("key/login.bin","key/mon.key")

        #! 1. データの準備(データファイルダウンロード・変換)
        log(["データファイルのダウンロード・変換を行ないます"])
        window.refresh()

        #! 1-1. Amazon注文データ
        try:
            log(["Amazonデータファイル変換"])
            window.refresh()
            FFC = FileFormatConversion()
            FFC.filename = values["input_file"]
            FFC.output_filename = f"amazon_{selected_date.year}{str(selected_date.month).zfill(2)}.csv"
            FFC.txt_to_csv()
        except FileNotFoundError:
            choice = sg.popup_ok_cancel('【エラー】Amazonファイルが選択されていません。\n'
                    'Amazonデータが必要ない場合は,"OK"を押してください。データ入力を開始します。\n'
                    'Amazonデータが必要の場合は、"NG"を押し,アプリを終了してください。\n'
                    '再起動を行ない、Amazonファイルを選択したうえでプログラムを実行してください。')
            # Check the user's selection and act accordingly
            if choice == 'OK':
                pass
            else:
                break

        #! 1-2. 楽天注文データ
        log(["楽天データダウンロード開始"])
        window.refresh()
        D1 = Download()
        D1.rakuten( get_config_value(login_data, "rakuten", "Common_ID"),
                    get_config_value(login_data, "rakuten", "Common_PW"),
                    get_config_value(login_data, "rakuten", "Personal_ID"),
                    get_config_value(login_data, "rakuten", "Personal_PW"),
                    get_config_value(login_data, "rakuten", "Download_ID"),
                    get_config_value(login_data, "rakuten", "Download_PW")
                    )

        #! 1-3. Yahoo注文データ
        log(["Yahooデータダウンロード開始"])
        window.refresh()
        D2 = Download()
        D2.yahoo(get_config_value(login_data, "yahoo", "Login_ID"),
                get_config_value(login_data, "yahoo", "Login_PW")
                )

        #! 1-4. 返品データ
        log(["返品数データダウンロード開始"])
        window.refresh()
        download_for_googledrive(saving_foldername="temp",search_list=["返品",str(selected_date.year).zfill(4),str(selected_date.month).zfill(2),".xlsx"],folder_id="1WszXFq7e-1yKvQ8gs45-09oWQYZZLeBM")

        #! 1-5.サンプル・不良・調整データ
        log(["サンプル・不良・調整データダウンロード開始"])
        window.refresh()
        download_for_googledrive(saving_foldername="temp",search_list=[f"サンプル・不良・調整_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}.xlsx"],folder_id="1RST89cKOd3U91EKOEvLATElfL5bQ7bqK")

        #! 1-6. 塗屋への発注データ
        log(["発注データダウンロード開始"])
        window.refresh()
        # googledriveから塗屋入庫データを取得
        download_for_googledrive(saving_foldername="temp",search_list=[f"発注_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}.xlsx"],folder_id="1uQULhUO5-LUGrAg9Ihy_KRCkkwHpR0We")

        #! 1-7. 塗屋からの入庫データ
        log(["入庫データダウンロード開始"])
        window.refresh()
        # googledriveから塗屋入庫データを取得
        download_for_googledrive(saving_foldername="temp",search_list=[f"山家_{str(selected_date.year).zfill(4)}-{str(selected_date.month).zfill(2)}",".csv"],folder_id="1-x9H9Ixt7tAKUnmQ0F3qilEHfVkuMWxU")

        #! 1-8. FBA納品数データ
        log(["FBA納品数データダウンロード開始"])
        window.refresh()
        download_for_googledrive(saving_foldername="temp",search_list=["FBA納品",str(selected_date.year).zfill(4),str(selected_date.month).zfill(2),".xlsx"],folder_id="1Nz2Gg6LABJp4pxiXY93VRS5NFSh54xVU")

        #! 2. データの取得（リスト化）
        #! 2-1. 出庫数取得
        log(["出庫数を取得します"])
        window.refresh()
        PP = Preparation()
        INPUT = Input()
        INPUT.amazon_issue_list, INPUT.amazonFBA_issue_list = PP.amazon_issue()
        INPUT.rakuten_issue_list = PP.rakuten_issue()
        INPUT.yahoo_issue_list = PP.yahoo_issue()

        #! 2-2. 返品取得
        log(["返品数を取得します"])
        window.refresh()
        INPUT.amazon_return_list,INPUT.amazonFBA_return_list,INPUT.rakuten_return_list,INPUT.yahoo_return_list = PP.returns()

        #! 2-3.サンプル・不良・調整数取得
        log(["サンプル・不良・調整数を取得します"])
        window.refresh()
        sample_list = PP.sample()
        defective_list = PP.defective()
        adjust_list = PP.adjust()

        #! 2-3. 塗屋発注数取得
        log(["発注数を取得します"])
        window.refresh()
        orders_list = PP.orders()

        #! 2-4. 塗屋入庫数取得
        log(["入庫数を取得します"])
        window.refresh()
        receipts_list = PP.receipts()

        #! 2-5. FBA納品数取得
        log(["FBA納品数を取得します"])
        window.refresh()
        deliveries_list = PP.deliveries()

        #! 3. 在庫表への入力
        log(['在庫表への入力を開始します'])
        # シリーズ毎に
        num = 0
        for series_name in series_list:
            num += 1
            print(f"--- {series_name}シリーズ ({num}/{len(series_list)}) ---")
            window.refresh()

            try:
                #! 3-0. 書き込むexcelファイルの準備
                file_name = f'data\\【新EC在庫表】{series_name}_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx'
                # excelをアプリで開く
                app = xw.App(visible=False)
                wb = app.books.open(f'{file_name}',read_only=False)
                # 計算を手動に設定
                app.calculation = 'manual'

                #! 3-1. 自社倉庫在庫表への入力
                log(['サンプル・不良数・調整数の入力'])
                window.refresh()
                INPUT.sample(sample_list) # サンプル数の入力
                INPUT.defective(defective_list) # 不良数の入力
                INPUT.adjust(adjust_list) # 調整数の入力

                #! 3-2. 出庫表への入力
                log(['出庫数・返品数の入力開始'])
                window.refresh()
                INPUT.single() # 単品の出庫数と返品数の入力
                INPUT.set() # セット商品の出庫数と返品数の入力
                INPUT.FBA() # セット商品のFBA出庫数と返品数の入力

                #! 3-3. 発注数・入庫数への入力
                # 塗屋発注数・塗屋入庫数入力
                log(['発注数・入庫数の入力開始'])
                window.refresh()
                INPUT.orders(orders_list)
                INPUT.receipts(receipts_list)

                app.calculation = 'automatic'
                for sheet in wb.sheets:
                    sheet.api.Calculate()  # 各シートの再計算
                # 計算を手動に設定
                app.calculation = 'manual'

                #! 3-4. FBA納品表への入力
                log(['FBA納品数の入力開始'])
                window.refresh()
                INPUT.deliveries(deliveries_list)

                #! 3-5. 書き込んだexcelファイルの保存
                # 計算を自動に戻して、全シートの再計算を実行
                app.calculation = 'automatic'
                for sheet in wb.sheets:
                    sheet.api.Calculate()  # 各シートの再計算

                #ファイルの保存とブックのクローズ
                wb.save(file_name)
                wb.close()
                log([f'ファイルの保存が完了しました。ファイル名「{file_name}」'])
                window.refresh()
                app.quit()
                time.sleep(5)

            except Exception as e:
                app.quit()
                log([f"エラー：{e}"])
                window.refresh()
                pass

        #! 4. sales_inventory.xlsx へ現自社倉庫在庫数とFBA倉庫在庫数、出庫数を記入
        log(["在庫数出庫数の履歴.xlsx の更新開始"])
        window.refresh()
        ADD = Add()
        ADD.inventory()
        log(["更新開始"])

        #! 5. 塗屋注残共有用ファイル作成
        log(["塗屋注残数共有ファイル作成開始"])
        # Naturalのみ
        copySheet(f'data\\【新EC在庫表】Natural_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx',\
                    f'data\\注残_{selected_date.year}年{str(selected_date.month).zfill(2)}月.xlsx',\
                        7)
        log(["作成完了"])

        #! 6. 注残リスト更新
        log(["自動更新用注残数リスト更新開始"])
        # 注残数取得（全シリーズの注残数を取得して一つの行列に格納）

        # 新規作成
        xw.App(visible=False)
        wb = xw.Book()
        ws = wb.sheets['Sheet1']

        num = 0
        for series_name in series_list:
            order_rest_list = []
            window.refresh()
            try:
                PP = Preparation()
                order_rest_list = PP.orderRest()
                # 入力
                for i in range(len(order_rest_list)):
                    ws.cells(1+num,1).value = order_rest_list[i][0]
                    ws.cells(1+num,2).value = order_rest_list[i][1]
                    ws.cells(1+num,3).value = order_rest_list[i][2]
                    ws.cells(1+num,4).value = int(order_rest_list[i][3])
                    num += 1
            except Exception as e:
                print(e)

        try:
            file_name = f"data\\（自動更新）注残数_{str(selected_date.year)}-{str(selected_date.month).zfill(2)}.xlsx"
            #ファイルの保存とブックのクローズ
            wb.save(file_name)
            wb.close()
            log(["ファイル保存完了"])
            window.refresh()
            time.sleep(5)
        except Exception as e:
            print("エラー：",e)
            window.refresh()
            pass

        log(["更新完了"])

        #! 7. GoogleDriveへのアップロード
        # 在庫表をアップロード
        log(["GoogleDriveへのアップロード開始"])
        window.refresh()
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["ECAIIS",f"{selected_date.year}年{selected_date.month}月在庫表"],
                            upload_folder_name="data",
                            upload_file_name_list=["在庫表",f"{selected_date.year}年{str(selected_date.month).zfill(2)}月"]
                            )
        # 山家様用注残共有用ファイルをアップロード
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["山家様用注残共有フォルダ"],
                            upload_folder_name="data",
                            upload_file_name_list=[f"注残_{str(selected_date.year)}年{str(selected_date.month).zfill(2)}月.xlsx"]
                            )
        # 自動更新用注残数ファイルをアップロード
        upload_to_googledrive(folder_id="1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO",
                            search_folder_name_list=["塗屋発注数データ"],
                            upload_folder_name="data",
                            upload_file_name_list=["自動更新",f"注残数_{str(selected_date.year)}-{str(selected_date.month).zfill(2)}.xlsx"]
                            )

        log(["アップロード完了"])
        window.refresh()

        log(["自動入力完了"])
        window.refresh()

    if event == "操作手順":
        sg.popup("1.日付を選択する。\n"
                "2.Amazon Seller Centralを開き、Amazonデータを準備する。\n"
                "3.「選択」ボタンからダイアログを開き、\n"
                "   ダウンロードしたAmazonファイルを選択する。\n"
                "4.「一月分」ボタンをクリック。"\
                ,no_titlebar=True)

    if event == "注意事項":
        sg.popup("・プログラム実行時にGoogleChromeが開かれていないこと。\n"
                "・プログラム実行中、Excelを開かないこと。\n"
                "・不具合等が発生した場合は本夛に連絡してください。\n"
                "    連絡先 : satoshi_honda@taiyo-chemicals.co.jp"\
                ,no_titlebar=True)

    if event == "概要（リンクを開く）":
        url = "https://www.notion.so/ECAIIS-2-2-2-2024-9-11-d6d808efcfa6400c93edbef2e691de8c?pvs=4"
        webbrowser.open(url)

    if event == "保存先フォルダを開く":
        url = "https://drive.google.com/drive/u/0/folders/1z79YIolCBdvEbV7YT5eAvsaSgg5D4LxO"
        webbrowser.open(url)

    if event == "製品情報":
        sg.popup(f"製品名: {PROGRAM_NAME}\n"\
                f"バージョン: ver.{PROGRAM_VERSION}\n"\
                f"更新日: {PROGRAM_UPDATE}",
                no_titlebar=True)